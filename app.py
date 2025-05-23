from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, make_response
import sqlite3
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import os
import csv
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io
import xlsxwriter
import pandas as pd
import shutil
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import Image

app = Flask(__name__)
app.secret_key = 'your_secret_key'

def get_db_connection():
    conn = sqlite3.connect("users.db")
    conn.row_factory = sqlite3.Row
    return conn

# Create admin_logs table if it doesn't exist
def create_admin_logs_table():
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS admin_logs (
            log_id INTEGER PRIMARY KEY AUTOINCREMENT,
            admin_id TEXT NOT NULL,
            action TEXT NOT NULL,
            timestamp DATETIME NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

# Create the admin logs table when the app starts
create_admin_logs_table()

@app.after_request
def add_header(response):
    """Add headers to both force latest IE rendering engine or Chrome Frame,
    and also to cache the rendered page for 10 minutes."""
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

def initialize_lab_computers():
    """Ensure all lab computers exist in the lab_computers table"""
    try:
        conn = get_db_connection()
        
        # First, check if the lab_computers table exists
        table_exists = conn.execute("""
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name='lab_computers'
        """).fetchone()
        
        # If table doesn't exist, create it
        if not table_exists:
            conn.execute("""
                CREATE TABLE lab_computers (
                    id INTEGER PRIMARY KEY,
                    lab_id INTEGER NOT NULL,
                    pc_number INTEGER NOT NULL,
                    status TEXT DEFAULT 'Available',
                    last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (lab_id) REFERENCES laboratories(lab_id),
                    UNIQUE(lab_id, pc_number)
                )
            """)
        
        # Get all labs and their capacities
        labs = conn.execute("SELECT lab_id, capacity FROM laboratories").fetchall()
        
        # Initialize PCs for each lab
        for lab in labs:
            lab_id = lab['lab_id']
            capacity = lab['capacity']
            
            # Check existing PCs for this lab
            existing_pcs = conn.execute("""
                SELECT pc_number FROM lab_computers
                WHERE lab_id = ?
            """, (lab_id,)).fetchall()
            
            existing_pc_numbers = [pc['pc_number'] for pc in existing_pcs]
            
            # Create missing PCs
            for i in range(1, capacity + 1):
                if i not in existing_pc_numbers:
                    conn.execute("""
                        INSERT INTO lab_computers (lab_id, pc_number, status, last_updated)
                        VALUES (?, ?, 'Available', CURRENT_TIMESTAMP)
                    """, (lab_id, i))
        
        # Update PC status based on active sessions
        active_sessions = conn.execute("""
            SELECT ls.session_id, ls.lab_id, r.computer_number
            FROM lab_sessions ls
            LEFT JOIN lab_reservations r ON ls.reservation_id = r.reservation_id
            WHERE ls.status = 'Active' AND r.computer_number IS NOT NULL
        """).fetchall()
        
        for session in active_sessions:
            lab_id = session['lab_id']
            pc_number = session['computer_number']
            
            # Update PC status to Used
            conn.execute("""
                UPDATE lab_computers
                SET status = 'Used',
                    last_updated = CURRENT_TIMESTAMP
                WHERE lab_id = ? AND pc_number = ?
            """, (lab_id, pc_number))
        
        conn.commit()
        conn.close()
        print("Lab computers initialized successfully")
        
    except sqlite3.Error as e:
        print(f"Database error initializing lab computers: {e}")
    
# Initialize lab computers when app starts
with app.app_context():
    initialize_lab_computers()

def format_time(value):
    try:
        # Try parsing as datetime first
        dt = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        return dt.strftime('%I:%M %p')
    except ValueError:
        try:
            # If that fails, try parsing as time
            if ':' in value:
                if len(value.split(':')) == 2:
                    dt = datetime.strptime(value, '%H:%M')
                else:
                    dt = datetime.strptime(value, '%H:%M:%S')
                return dt.strftime('%I:%M %p')
        except ValueError:
            # If all parsing fails, return the original value
            return value

def format_date(value):
    try:
        # Handle datetime string with timezone info (T)
        if 'T' in value:
            dt = datetime.strptime(value, '%Y-%m-%dT%H:%M')
        else:
            dt = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        # Format date and time separately
        date_str = dt.strftime('%B %d, %Y')
        time_str = dt.strftime('%I:%M %p')
        return date_str, time_str
    except ValueError:
        return value, value

app.jinja_env.filters['format_time'] = format_time
app.jinja_env.filters['format_date'] = format_date


@app.route('/')
def home():

    if 'user' in session:
        flash("Logged in successfully.", "success")
        return render_template('dashboard.html')
    else:
        return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    email = request.form.get('email')
    password = request.form.get('pass')
    conn = get_db_connection()
    
    try:
        user = conn.execute("SELECT * FROM users WHERE email = ? AND password = ?", (email, password)).fetchone()
        
        if user:
            if user['role'] == 'admin':
                session['user'] = 'admin'
                session['is_admin'] = True
                flash("Logged in as admin.", "success")
                return redirect(url_for('admin_dashboard'))
            else:
                student = conn.execute("SELECT * FROM students WHERE idno = ?", (email,)).fetchone()
                if student:
                    session['user'] = student['idno']
                session['is_admin'] = False
                flash("Logged in successfully.", "success")
                return redirect(url_for('dashboard'))
        else:
            flash("Student record not found.", "danger")
            return redirect(url_for('home'))
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('home'))
    finally:
        conn.close()



@app.route('/logout')
def logout():
    session.pop('user', None)
    flash("You have been logged out.", "info")
    return redirect(url_for('home'))

@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        flash("Please log in to access the dashboard.", "warning")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    
    # 1) Retrieve the student's course and other info
    student = conn.execute("SELECT * FROM students WHERE idno = ?", (session['user'],)).fetchone()
    if not student:
        conn.close()
        flash("Student record not found.", "danger")
        return redirect(url_for('home'))
    
    # 2) Determine max sessions based on course
    if student['course'] in ("BSIT", "BSCS"):
        max_sessions = 30
    else:
        max_sessions = 15

    # 3) Count completed lab sessions
    used_sessions = conn.execute("""
        SELECT COUNT(*) as c 
        FROM lab_sessions 
        WHERE student_id = ? AND status = 'Completed'
    """, (session['user'],)).fetchone()['c']
    
    # 4) Calculate remaining sessions
    remaining_sessions = max_sessions - used_sessions
    if remaining_sessions < 0:
        remaining_sessions = 0
        
    # 5) Calculate behavior points and free sessions
    behavior_points = conn.execute("""
        SELECT COALESCE(SUM(behavior_points), 0) as total
        FROM lab_sessions
        WHERE student_id = ? AND behavior_points > 0
    """, (session['user'],)).fetchone()['total']
    
    free_sessions = behavior_points // 3
    points_until_next = 3 - (behavior_points % 3) if behavior_points % 3 > 0 else 0
    
    # Add free sessions to remaining sessions
    if free_sessions > 0:
        remaining_sessions += free_sessions

    # 6) Get active announcements
    announcements = conn.execute("""
        SELECT * FROM announcements 
        WHERE is_active = 1 
        AND (expiry_date IS NULL OR expiry_date >= datetime('now'))
        ORDER BY posted_date DESC
    """).fetchall()

    # 7) Get the next upcoming reservation
    upcoming_reservation = conn.execute("""
        SELECT lr.*, l.room_number, l.building
        FROM lab_reservations lr
        JOIN laboratories l ON lr.lab_id = l.lab_id
        WHERE lr.student_id = ? 
        AND lr.reservation_date >= date('now')
        AND lr.status = 'Approved'
        ORDER BY lr.reservation_date, lr.start_time 
        LIMIT 1
    """, (session['user'],)).fetchone()
    
    # 8) Get active session if any
    active_session = conn.execute("""
        SELECT ls.*, l.room_number, l.building
        FROM lab_sessions ls
        JOIN laboratories l ON ls.lab_id = l.lab_id
        WHERE ls.student_id = ? AND ls.status = 'Active'
        LIMIT 1
    """, (session['user'],)).fetchone()
    
    conn.close()
    
    return render_template(
        'dashboard.html',
        student=student,
        reservation=upcoming_reservation,
        active_session=active_session,
        remaining_sessions=remaining_sessions,
        announcements=announcements,
        behavior_points=behavior_points,
        free_sessions=free_sessions,
        points_until_next=points_until_next
    )



@app.route('/profile')
def profile():
    if 'user' not in session:
        flash("Please log in to access your profile.", "warning")
        return redirect(url_for('home'))

    conn = get_db_connection()
    student = conn.execute("SELECT * FROM students WHERE idno = ?", (session['user'],)).fetchone()
    conn.close()

    if not student:
        flash("Student information not found.", "danger")
        return redirect(url_for('dashboard'))
    
    return render_template('profile.html', student=student)


@app.route('/edit_profile', methods=['GET', 'POST'])
def edit_profile():
    if 'user' not in session:
        flash("Please log in to access your profile.", "warning")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    
    if request.method == 'POST':
        # Retrieve form data
        idno = request.form.get('idno')
        lastname = request.form.get('lastname')
        firstname = request.form.get('firstname')
        midname = request.form.get('midname')
        course = request.form.get('course')
        year_level = request.form.get('year_level')
        email_address = request.form.get('email_address')
        delete_image = request.form.get('delete_image') == 'yes'  # Check if delete checkbox is checked

        # Check for email conflicts
        existing_email = conn.execute(
            "SELECT * FROM students WHERE email_address = ? AND idno != ?",
            (email_address, idno)
        ).fetchone()
        if existing_email:
            flash("Email is already used by another account.", "danger")
            conn.close()
            return redirect(url_for('edit_profile'))
        
        # Get current student info to access current image path
        current_student = conn.execute("SELECT image_path FROM students WHERE idno = ?", (idno,)).fetchone()
        image_path = current_student['image_path']  # Keep current path by default
        
        # Handle profile image upload or deletion
        profile_image = request.files.get('profile_image')
        
        if delete_image and image_path:
            # Delete the physical file if it exists
            if image_path:
                file_path = os.path.join('static', image_path)
                if os.path.exists(file_path):
                    os.remove(file_path)
            # Set image_path to None in the database
            image_path = None
            
        elif profile_image and profile_image.filename:
            # Delete old image file if exists
            if image_path:
                old_file_path = os.path.join('static', image_path)
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)
                    
            # Process new image
            filename = secure_filename(profile_image.filename)
            upload_folder = os.path.join('static', 'uploads')
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)
            save_path = os.path.join(upload_folder, filename)
            profile_image.save(save_path)
            image_path = os.path.join('uploads', filename).replace(os.sep, '/')

        # Update student record including image path
        conn.execute("""
            UPDATE students
            SET lastname = ?,
                firstname = ?,
                midname = ?,
                course = ?,
                year_level = ?,
                email_address = ?,
                image_path = ?
            WHERE idno = ?
        """, (lastname, firstname, midname, course, year_level, email_address, image_path, idno))
        
        conn.commit()
        conn.close()
        flash("Profile updated successfully!", "success")
        return redirect(url_for('profile'))
    
    else:
        # For GET requests, retrieve the current student info
        student = conn.execute("SELECT * FROM students WHERE idno = ?", (session['user'],)).fetchone()
        conn.close()
        if not student:
            flash("Student information not found.", "danger")
            return redirect(url_for('dashboard'))
        return render_template('edit_profile.html', student=student)    



@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        try:
            idno = request.form.get('idno')
            lastname = request.form.get('lastname')
            firstname = request.form.get('firstname')
            midname = request.form.get('midname')
            course = request.form.get('course')
            year_level = request.form.get('year_level')  # This will now get the text format (e.g., "1st Year")
            email = request.form.get('email')
            password = request.form.get('password')

            # Check that all required fields are provided
            if not all([idno, lastname, firstname, course, year_level, email, password]):
                flash("Missing required fields", "danger")
                return redirect(url_for('register'))
            
            conn = get_db_connection()

            # Check if the student ID is already registered
            existing_student = conn.execute("SELECT * FROM students WHERE idno = ?", (idno,)).fetchone()
            if existing_student:
                flash("Student ID is already used by another account.", "danger")
                conn.close()
                return redirect(url_for('register'))

            # Check if the email is already used
            existing_email = conn.execute("SELECT * FROM students WHERE email_address = ?", (email,)).fetchone()
            if existing_email:
                flash("Email is already used by another account.", "danger")
                conn.close()
                return redirect(url_for('register'))

            # Begin transaction
            conn.execute("BEGIN TRANSACTION")
            
            try:
                # Insert into users table first with idno as email/username
                conn.execute("""
                    INSERT INTO users (email, password, role) 
                    VALUES (?, ?, 'student')
                """, (idno, password))

                # Then insert into students table with year_level as text
                conn.execute("""
                    INSERT INTO students (idno, lastname, firstname, midname, course, year_level, email_address) 
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (idno, lastname, firstname, midname, course, year_level, email))
                
                # Commit the transaction
                conn.commit()
                flash("Registration successful! You can now log in.", "success")
                return redirect(url_for('home'))
                
            except sqlite3.Error as e:
                # Roll back in case of error
                conn.execute("ROLLBACK")
                flash(f"Database error: {str(e)}", "danger")
                return redirect(url_for('register'))

        except Exception as e:
            flash(f"Error: {str(e)}", "danger")
            return redirect(url_for('register'))
        
        finally:
            conn.close()

    return render_template('register.html')




@app.route('/remaining_sessions')
def remaining_sessions():
    if 'user' not in session:
        flash("Please log in to view remaining sessions.", "warning")
        return redirect(url_for('home'))

    conn = get_db_connection()

    # 1) Retrieve the student's course
    student = conn.execute(
        "SELECT course FROM students WHERE idno = ?",
        (session['user'],)
    ).fetchone()

    if not student:
        conn.close()
        flash("Student record not found.", "danger")
        return redirect(url_for('home'))

    # 2) Determine allocated sessions based on the course
    if student['course'] in ("BSIT", "BSCS"):
        allocated_sessions = 30
    else:
        allocated_sessions = 15

    # 3) Count how many sessions the student has used
    used_sessions = conn.execute("""
        SELECT COUNT(*) as c 
        FROM lab_sessions 
        WHERE student_id = ? AND status = 'Completed'
    """, (session['user'],)).fetchone()['c']

    # 4) Calculate remaining sessions
    remaining_sessions = allocated_sessions - used_sessions
    if remaining_sessions < 0:
        remaining_sessions = 0  # clamp to 0 if they exceed
        
    # 5) Calculate behavior points and free sessions
    behavior_points = conn.execute("""
        SELECT COALESCE(SUM(behavior_points), 0) as total
        FROM lab_sessions
        WHERE student_id = ? AND behavior_points > 0
    """, (session['user'],)).fetchone()['total']
    
    free_sessions = behavior_points // 3
    points_until_next = 3 - (behavior_points % 3) if behavior_points % 3 > 0 else 0
    
    # Add free sessions to remaining sessions
    if free_sessions > 0:
        remaining_sessions += free_sessions

    # 6) Retrieve an upcoming reservation (if any)
    reservation = conn.execute("""
        SELECT lr.*, l.room_number, l.building
        FROM lab_reservations lr
        JOIN laboratories l ON lr.lab_id = l.lab_id
        WHERE lr.student_id = ? 
        AND lr.reservation_date >= date('now')
        AND lr.status = 'Approved'
        ORDER BY lr.reservation_date, lr.start_time 
        LIMIT 1
    """, (session['user'],)).fetchone()

    # 7) Retrieve lab session history
    sessions = conn.execute("""
        SELECT 
            ls.*,
            l.room_number,
            l.building,
            ROUND(CAST((julianday(ls.check_out_time) - julianday(ls.check_in_time)) * 24 AS REAL), 2) as duration
        FROM lab_sessions ls
        JOIN laboratories l ON ls.lab_id = l.lab_id
        WHERE ls.student_id = ? AND ls.status = 'Completed'
        ORDER BY ls.check_in_time DESC
    """, (session['user'],)).fetchall()

    conn.close()

    return render_template(
        'remaining_sessions.html',
        remaining_sessions=remaining_sessions,
        reservation=reservation,
        sessions=sessions,
        behavior_points=behavior_points,
        free_sessions=free_sessions,
        points_until_next=points_until_next
    )



@app.route('/sit_in_history')
def sit_in_history():
    if 'user' not in session:
        flash("Please log in to view your sit-in history.", "warning")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    
    # Get all completed lab sessions with location details and feedback status
    sessions = conn.execute("""
        SELECT 
            ls.*,
            l.room_number,
            l.building,
            ROUND(CAST((julianday(ls.check_out_time) - julianday(ls.check_in_time)) * 24 AS REAL), 2) as duration,
            CASE WHEN sf.feedback_id IS NOT NULL THEN 1 ELSE 0 END as has_feedback
        FROM lab_sessions ls
        JOIN laboratories l ON ls.lab_id = l.lab_id
        LEFT JOIN session_feedback sf ON ls.session_id = sf.session_id AND sf.student_id = ls.student_id
        WHERE ls.student_id = ? AND ls.status = 'Completed'
        ORDER BY ls.check_in_time DESC
    """, (session['user'],)).fetchall()
    
    conn.close()
    
    return render_template('sit_in_history.html', sessions=sessions)




@app.route('/reservation', methods=['GET', 'POST'])
def reservation():
    if 'user' not in session:
        flash("Please log in to make a reservation.", "warning")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    
    if request.method == 'POST':
        lab_id = request.form.get('lab_id')
        reservation_date = request.form.get('reservation_date')
        start_time = request.form.get('start_time')
        end_time = request.form.get('end_time')
        purpose = request.form.get('purpose')
        student_id = session['user']
        computer_number = request.form.get('computer_number')  # Get the computer number
        
        try:
            # Check if end time is after start time
            start = datetime.strptime(start_time, '%H:%M')
            end = datetime.strptime(end_time, '%H:%M')
            if end <= start:
                flash("End time must be after start time.", "danger")
                return redirect(url_for('reservation'))
            
            # Check for overlapping reservations
            overlapping = conn.execute("""
                SELECT * FROM lab_reservations 
                WHERE lab_id = ? 
                AND reservation_date = ? 
                AND status = 'Approved'
                AND (
                    (start_time <= ? AND end_time > ?) OR
                    (start_time < ? AND end_time >= ?) OR
                    (start_time >= ? AND end_time <= ?)
                )
            """, (lab_id, reservation_date, start_time, start_time, 
                  end_time, end_time, start_time, end_time)).fetchone()
            
            if overlapping:
                flash("This time slot is already reserved.", "danger")
                return redirect(url_for('reservation'))
            
            # Create the reservation
            conn.execute("""
                    INSERT INTO lab_reservations (student_id, lab_id, reservation_date, start_time, end_time, purpose, status, computer_number)
                    VALUES (?, ?, ?, ?, ?, ?, 'Pending', ?)
                """, (student_id, lab_id, reservation_date, start_time, end_time, purpose, computer_number))
            conn.commit()
            flash("Reservation submitted successfully!", "success")
                
        except ValueError as e:
            flash(f"Invalid time format: {str(e)}", "danger")
        except sqlite3.Error as e:
            flash(f"Database error: {str(e)}", "danger")
        
        return redirect(url_for('reservation'))
    
    # For GET request, retrieve available labs and upcoming reservations
    try:
        # Get all labs
        labs = conn.execute("SELECT * FROM laboratories ORDER BY building, room_number").fetchall()
        
        # Get upcoming reservations for the student
        reservations = conn.execute("""
                SELECT lr.*, l.building, l.room_number, 
                    CASE 
                        WHEN lr.status = 'Approved' THEN 'badge-success'
                        WHEN lr.status = 'Pending' THEN 'badge-warning'
                        WHEN lr.status = 'Rejected' THEN 'badge-danger'
                        ELSE 'badge-info'
                    END as status_badge,
                    lr.rejection_reason,
                    lr.computer_number
                FROM lab_reservations lr
                JOIN laboratories l ON lr.lab_id = l.lab_id
                WHERE lr.student_id = ? 
                AND (lr.status = 'Pending' OR 
                     lr.status = 'Rejected' OR
                    (lr.status = 'Approved' AND lr.reservation_date >= date('now')))
                ORDER BY lr.reservation_date ASC, lr.start_time ASC
        """, (session['user'],)).fetchall()
        
        # Get student info
        student = conn.execute("SELECT * FROM students WHERE idno = ?", (session['user'],)).fetchone()
        # Calculate remaining sessions (reuse dashboard logic)
        if student['course'] in ("BSIT", "BSCS"):
            max_sessions = 30
        else:
            max_sessions = 15
        used_sessions = conn.execute("""
            SELECT COUNT(*) as c 
            FROM lab_sessions 
            WHERE student_id = ? AND status = 'Completed'
        """, (session['user'],)).fetchone()['c']
        remaining_sessions = max_sessions - used_sessions
        if remaining_sessions < 0:
            remaining_sessions = 0
        behavior_points = conn.execute("""
            SELECT COALESCE(SUM(behavior_points), 0) as total
            FROM lab_sessions
            WHERE student_id = ? AND behavior_points > 0
        """, (session['user'],)).fetchone()['total']
        free_sessions = behavior_points // 3
        if free_sessions > 0:
            remaining_sessions += free_sessions
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        labs = []
        reservations = []
        student = None
        remaining_sessions = 0
    conn.close()
    return render_template('reservation.html', 
                         labs=labs, 
                         reservations=reservations,
                         now=datetime.now().strftime('%Y-%m-%d'),
                         student=student,
                         remaining_sessions=remaining_sessions)

@app.route('/admin')
def admin_dashboard():
    # Check if user is logged in
    if 'user' not in session:
        flash("Please log in to access the admin panel.", "warning")
        return redirect(url_for('home'))
    
    # Check if user is admin
    if not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('dashboard'))
    
    try:
        # Get database connection
        conn = get_db_connection()
        
        # Get student statistics
        total_students = conn.execute("SELECT COUNT(*) as count FROM students").fetchone()['count']
        active_sessions = conn.execute("SELECT COUNT(*) as count FROM lab_sessions WHERE status = 'Active'").fetchone()['count']
        total_sit_ins = conn.execute("SELECT COUNT(*) as count FROM lab_sessions WHERE status = 'Completed'").fetchone()['count']
        
        # Fetch all students with their details
        students_basic = conn.execute("""
            SELECT idno, lastname, firstname, midname, course, year_level, email_address, free_sessions_used 
            FROM students 
            ORDER BY lastname, firstname
        """).fetchall()
        
        # Convert to list of dictionaries for modification
        students = []
        for student in students_basic:
            # Create a mutable dictionary from the row
            student_dict = dict(student)
            
            # Get used sessions count (Completed + Active)
            used_sessions = conn.execute("""
                SELECT COUNT(*) as count 
                FROM lab_sessions 
                WHERE student_id = ? AND (status = 'Completed' OR status = 'Active')
            """, (student['idno'],)).fetchone()['count']
            student_dict['used_sessions'] = used_sessions
            
            # Get total behavior points
            behavior_points = conn.execute("""
                SELECT COALESCE(SUM(behavior_points), 0) as total_points
                FROM lab_sessions 
                WHERE student_id = ? AND behavior_points > 0
            """, (student['idno'],)).fetchone()['total_points']
            student_dict['behavior_points'] = behavior_points
            
            # Calculate remaining sessions using new logic
            max_sessions = 30 if student['course'] in ("BSIT", "BSCS") else 15
            free_sessions_earned = behavior_points // 3
            free_sessions_used = student['free_sessions_used'] if student['free_sessions_used'] is not None else 0
            free_sessions_available = free_sessions_earned - free_sessions_used
            if free_sessions_available < 0:
                free_sessions_available = 0
            base_remaining = max_sessions - used_sessions
            if base_remaining < 0:
                base_remaining = 0
            total_remaining = base_remaining + free_sessions_available
            student_dict['free_sessions_available'] = free_sessions_available
            student_dict['remaining_sessions'] = total_remaining
            students.append(student_dict)
        
        # Fetch all available laboratories
        labs = conn.execute("""
            SELECT lab_id, room_number, building 
            FROM laboratories 
            WHERE status = 'Available'
            ORDER BY building, room_number
        """).fetchall()
        
        # Close the connection
        conn.close()

        # Render admin template with student and lab data
        return render_template('admin.html', 
                             students=students, 
                             labs=labs,
                             total_students=total_students,
                             active_sessions=active_sessions,
                             total_sit_ins=total_sit_ins)
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('home'))

@app.route('/admin/announcements', methods=['GET', 'POST'])
def admin_announcements():
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    
    if request.method == 'POST':
        title = request.form.get('title')
        content = request.form.get('content')
        expiry_date = request.form.get('expiry_date')
        
        try:
            # Get current time in Philippines timezone
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            conn.execute("""
                INSERT INTO announcements (title, content, posted_by, posted_date, expiry_date, is_active)
                VALUES (?, ?, ?, ?, ?, 1)
            """, (title, content, 1, current_time, expiry_date))  # Using 1 as admin's user ID
            conn.commit()
            flash("Announcement posted successfully!", "success")
        except sqlite3.Error as e:
            flash(f"Error posting announcement: {str(e)}", "danger")
        
        return redirect(url_for('admin_announcements'))
    
    try:
        # First, update expired announcements to inactive
        conn.execute("""
            UPDATE announcements 
            SET is_active = 0 
            WHERE expiry_date IS NOT NULL 
            AND expiry_date < datetime('now')
            AND is_active = 1
        """)
        conn.commit()
        
        # Then fetch all announcements for admin view
        announcements = conn.execute("""
            SELECT *,
                CASE 
                    WHEN expiry_date IS NOT NULL AND expiry_date < datetime('now') THEN 0
                    ELSE is_active 
                END as display_status,
                CASE
                    WHEN expiry_date IS NOT NULL AND expiry_date < datetime('now') THEN 1
                    ELSE 0
                END as is_expired
            FROM announcements 
            ORDER BY posted_date DESC
        """).fetchall()
        
        return render_template('admin_announcements.html', announcements=announcements)
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_dashboard'))
    finally:
        conn.close()

@app.route('/admin/announcements/toggle/<int:announcement_id>', methods=['GET', 'POST'])
def toggle_announcement(announcement_id):
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    try:
        # Get current announcement status
        current = conn.execute(
            "SELECT is_active FROM announcements WHERE announcement_id = ?", 
            (announcement_id,)
        ).fetchone()
        
        if request.method == 'POST' and not current['is_active']:
            # Activating with new expiry date
            expiry_date = request.form.get('expiry_date')
            conn.execute("""
                UPDATE announcements 
                SET is_active = 1, expiry_date = ?
                WHERE announcement_id = ?
            """, (expiry_date, announcement_id))
        else:
            # Simple toggle (deactivation)
            conn.execute("""
                UPDATE announcements 
                SET is_active = CASE WHEN is_active = 1 THEN 0 ELSE 1 END 
                WHERE announcement_id = ?
            """, (announcement_id,))
        
        conn.commit()
        flash("Announcement status updated successfully!", "success")
    except sqlite3.Error as e:
        flash(f"Error updating announcement: {str(e)}", "danger")
    finally:
        conn.close()
    
    return redirect(url_for('admin_announcements'))

@app.route('/admin/announcements/delete/<int:announcement_id>')
def delete_announcement(announcement_id):
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM announcements WHERE announcement_id = ?", (announcement_id,))
        conn.commit()
        flash("Announcement deleted successfully!", "success")
    except sqlite3.Error as e:
        flash(f"Error deleting announcement: {str(e)}", "danger")
    finally:
        conn.close()
    
    return redirect(url_for('admin_announcements'))

@app.route('/api/student_sessions/<student_id>')
def get_student_sessions(student_id):
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    conn = get_db_connection()
    try:
        # Get student's course and free_sessions_used
        student = conn.execute("SELECT course, free_sessions_used FROM students WHERE idno = ?", (student_id,)).fetchone()
        if not student:
            return jsonify({'error': 'Student not found'}), 404

        max_sessions = 30 if student['course'] in ("BSIT", "BSCS") else 15
        free_sessions_used = student['free_sessions_used'] if student['free_sessions_used'] is not None else 0

        # Count used sessions (Completed + Active)
        used_sessions = conn.execute("""
            SELECT COUNT(*) as count 
            FROM lab_sessions 
            WHERE student_id = ? AND (status = 'Completed' OR status = 'Active')
        """, (student_id,)).fetchone()['count']

        # Calculate behavior points and free sessions
        behavior_points = conn.execute("""
            SELECT COALESCE(SUM(behavior_points), 0) as total
            FROM lab_sessions
            WHERE student_id = ? AND behavior_points > 0
        """, (student_id,)).fetchone()['total']
        free_sessions_earned = behavior_points // 3
        free_sessions_available = free_sessions_earned - free_sessions_used
        if free_sessions_available < 0:
            free_sessions_available = 0

        base_remaining = max_sessions - used_sessions
        if base_remaining < 0:
            base_remaining = 0
        total_remaining = base_remaining + free_sessions_available

        return jsonify({
            'remaining_sessions': total_remaining
        })

    except sqlite3.Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/start_session', methods=['POST'])
def start_session():
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json()
    student_id = data.get('student_id')
    lab_id = data.get('lab_id')
    purpose = data.get('purpose')

    if not all([student_id, lab_id, purpose]):
        return jsonify({'error': 'Missing required fields'}), 400

    conn = get_db_connection()
    try:
        # Check if student has any active sessions
        active_session = conn.execute("""
            SELECT * FROM lab_sessions 
            WHERE student_id = ? AND status = 'Active'
        """, (student_id,)).fetchone()

        if active_session:
            return jsonify({
                'success': False,
                'message': 'Student already has an active session'
            }), 400

        # Get student info
        student = conn.execute("SELECT course, free_sessions_used FROM students WHERE idno = ?", (student_id,)).fetchone()
        max_sessions = 30 if student['course'] in ("BSIT", "BSCS") else 15
        free_sessions_used = student['free_sessions_used'] if student['free_sessions_used'] is not None else 0

        # Count used sessions (Completed + Active)
        used_sessions = conn.execute("""
            SELECT COUNT(*) as count 
            FROM lab_sessions 
            WHERE student_id = ? AND (status = 'Completed' OR status = 'Active')
        """, (student_id,)).fetchone()['count']

        # Calculate behavior points and free sessions
        behavior_points = conn.execute("""
            SELECT COALESCE(SUM(behavior_points), 0) as total
            FROM lab_sessions
            WHERE student_id = ? AND behavior_points > 0
        """, (student_id,)).fetchone()['total']
        free_sessions_earned = behavior_points // 3
        free_sessions_available = free_sessions_earned - free_sessions_used
        if free_sessions_available < 0:
            free_sessions_available = 0

        base_remaining = max_sessions - used_sessions
        if base_remaining < 0:
            base_remaining = 0
        total_remaining = base_remaining + free_sessions_available

        if total_remaining <= 0:
            return jsonify({
                'success': False,
                'message': 'No remaining sessions available'
            }), 400

        # If no base sessions left, use a free session
        if base_remaining == 0 and free_sessions_available > 0:
            conn.execute("""
                UPDATE students SET free_sessions_used = free_sessions_used + 1 WHERE idno = ?
            """, (student_id,))

        # Get current time in Philippines timezone
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Start new session with current time
        conn.execute("""
            INSERT INTO lab_sessions (
                student_id, lab_id, check_in_time, status, notes, created_by
            ) VALUES (?, ?, ?, 'Active', ?, ?)
        """, (student_id, lab_id, current_time, purpose, session['user']))
        
        conn.commit()
        return jsonify({'success': True})

    except sqlite3.Error as e:
        return jsonify({
            'success': False,
            'message': f"Database error: {str(e)}"
        }), 500
    finally:
        conn.close()

@app.route('/admin/current-sit-in')
def admin_current_sit_in():
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        
        # Fetch all active sessions with student details
        active_sessions = conn.execute("""
            SELECT 
                ls.session_id,
                s.idno,
                s.lastname,
                s.firstname,
                s.midname,
                s.course,
                s.year_level,
                ls.check_in_time,
                ls.notes as purpose,
                l.room_number,
                l.building,
                ls.behavior_points
            FROM lab_sessions ls
            JOIN students s ON ls.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE ls.status = 'Active'
            ORDER BY ls.check_in_time DESC
        """).fetchall()
        
        conn.close()
        
        return render_template('admin_current_sit_in.html', active_sessions=active_sessions)
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/current-sit-in/logout/<student_id>')
def logout_session(student_id):
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    try:
        # Get current time in Philippines timezone
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Get the lab session details before updating it
        session_details = conn.execute("""
            SELECT ls.lab_id, r.computer_number
            FROM lab_sessions ls
            LEFT JOIN lab_reservations r ON ls.reservation_id = r.reservation_id
            WHERE ls.student_id = ? AND ls.status = 'Active'
        """, (student_id,)).fetchone()
        
        # Update the session to completed
        conn.execute("""
            UPDATE lab_sessions 
            SET status = 'Completed',
                check_out_time = ?
            WHERE student_id = ? AND status = 'Active'
        """, (current_time, student_id))
        
        # If this session has an associated PC, update its status to Available
        if session_details and session_details['computer_number']:
            lab_id = session_details['lab_id']
            pc_number = session_details['computer_number']
            
            # Update the PC status to Available
            conn.execute("""
                UPDATE lab_computers
                SET status = 'Available',
                    last_updated = ?
                WHERE lab_id = ? AND pc_number = ?
            """, (current_time, lab_id, pc_number))
            
            # Log that the PC was freed
            lab_info = conn.execute("""
                SELECT building, room_number FROM laboratories WHERE lab_id = ?
            """, (lab_id,)).fetchone()
            
            if lab_info:
                location = f"{lab_info['building']} - Room {lab_info['room_number']}"
                conn.execute("""
                    INSERT INTO admin_logs (admin_id, action, timestamp)
                    VALUES (?, ?, datetime('now'))
                """, (session['user'], f"PC #{pc_number} in {location} freed after student {student_id} was logged out"))
        
        conn.commit()
    except sqlite3.Error as e:
        flash(f"Error logging out student: {str(e)}", "danger")
    finally:
        conn.close()
    
    return redirect(url_for('admin_current_sit_in'))

@app.route('/admin/sit-in-records')
def admin_sit_in_records():
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        
        # Get purpose/language statistics
        purpose_stats = conn.execute("""
            SELECT 
                notes as purpose,
                COUNT(*) as count
            FROM lab_sessions
            WHERE status = 'Completed'
            GROUP BY notes
            ORDER BY count DESC
        """).fetchall()
        
        # Get laboratory usage statistics
        lab_stats = conn.execute("""
            SELECT 
                l.building,
                l.room_number,
                COUNT(*) as count
            FROM lab_sessions ls
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE ls.status = 'Completed'
            GROUP BY l.building, l.room_number
            ORDER BY count DESC
        """).fetchall()
        
        # Get total completed sessions
        total_sessions = conn.execute("""
            SELECT COUNT(*) as total
            FROM lab_sessions
            WHERE status = 'Completed'
        """).fetchone()['total']
        
        # Fetch all completed sessions with student details
        completed_sessions = conn.execute("""
            SELECT 
                ls.session_id,
                s.idno,
                s.lastname,
                s.firstname,
                s.midname,
                s.course,
                s.year_level,
                DATE(ls.check_in_time) as date,
                ls.check_in_time,
                ls.check_out_time,
                ls.notes as purpose,
                l.room_number,
                l.building,
                ls.behavior_points,
                ROUND(CAST((julianday(ls.check_out_time) - julianday(ls.check_in_time)) * 24 AS REAL), 2) as duration
            FROM lab_sessions ls
            JOIN students s ON ls.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE ls.status = 'Completed'
            ORDER BY ls.check_in_time DESC
        """).fetchall()
        
        conn.close()
        
        return render_template('admin_sit_in_records.html', 
                             completed_sessions=completed_sessions,
                             purpose_stats=purpose_stats,
                             lab_stats=lab_stats,
                             total_sessions=total_sessions)
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/reports')
def admin_reports():
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    try:
        # Get all students for the dropdown
        students = conn.execute("""
            SELECT idno, lastname, firstname, midname, course, year_level, email_address
            FROM students
            ORDER BY lastname, firstname
        """).fetchall()
        
        # Get selected student's data if student_id is provided
        selected_student = None
        student_sessions = []
        student_feedback = []
        
        student_id = request.args.get('student_id')
        if student_id:
            # Get student details
            selected_student = conn.execute("""
                SELECT idno, lastname, firstname, midname, course, year_level, email_address
                FROM students
                WHERE idno = ?
            """, (student_id,)).fetchone()
            
            # Get student's sit-in sessions
            student_sessions = conn.execute("""
                SELECT 
                    s.idno,
                    s.lastname,
                    s.firstname,
                    s.midname,
                    ls.notes as purpose,
                    l.room_number,
                    l.building,
                    ls.check_in_time,
                    ls.check_out_time,
                    DATE(ls.check_in_time) as date,
                    ROUND(CAST((julianday(ls.check_out_time) - julianday(ls.check_in_time)) * 24 AS REAL), 2) as duration
                FROM lab_sessions ls
                JOIN students s ON ls.student_id = s.idno
                JOIN laboratories l ON ls.lab_id = l.lab_id
                WHERE ls.student_id = ? AND ls.status = 'Completed'
                ORDER BY ls.check_in_time DESC
            """, (student_id,)).fetchall()
            
            # Get student's feedback
            student_feedback = conn.execute("""
                SELECT 
                    s.idno,
                    s.lastname,
                    s.firstname,
                    s.midname,
                    l.room_number,
                    l.building,
                    DATE(ls.check_in_time) as date,
                    sf.rating,
                    sf.comments
                FROM session_feedback sf
                JOIN lab_sessions ls ON sf.session_id = ls.session_id
                JOIN students s ON sf.student_id = s.idno
                JOIN laboratories l ON ls.lab_id = l.lab_id
                WHERE sf.student_id = ?
                ORDER BY ls.check_in_time DESC
            """, (student_id,)).fetchall()
        
        return render_template(
            "admin_reports.html",
            students=students,
            selected_student=selected_student,
            student_sessions=student_sessions,
            student_feedback=student_feedback
        )
    finally:
        conn.close()

@app.route('/submit_feedback/<int:session_id>', methods=['GET', 'POST'])
def submit_feedback(session_id):
    if 'user' not in session:
        flash("Please log in to submit feedback.", "warning")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    try:
        # Check if the session exists and belongs to the student
        session_data = conn.execute("""
            SELECT 
                ls.*, 
                l.room_number, 
                l.building,
                ROUND(CAST((julianday(ls.check_out_time) - julianday(ls.check_in_time)) * 24 AS REAL), 2) as duration
            FROM lab_sessions ls
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE ls.session_id = ? AND ls.student_id = ? AND ls.status = 'Completed'
        """, (session_id, session['user'])).fetchone()
        
        if not session_data:
            flash("Session not found or you don't have permission to submit feedback.", "danger")
            return redirect(url_for('sit_in_history'))
        
        # Check if feedback already exists
        existing_feedback = conn.execute("""
            SELECT * FROM session_feedback 
            WHERE session_id = ? AND student_id = ?
        """, (session_id, session['user'])).fetchone()
        
        if existing_feedback:
            flash("You have already submitted feedback for this session.", "warning")
            return redirect(url_for('sit_in_history'))
        
        if request.method == 'POST':
            rating = request.form.get('rating')
            comments = request.form.get('comments')
            
            if not rating:
                flash("Please provide a rating.", "danger")
                return redirect(url_for('submit_feedback', session_id=session_id))
            
            # Insert the feedback
            conn.execute("""
                INSERT INTO session_feedback (session_id, student_id, rating, comments)
                VALUES (?, ?, ?, ?)
            """, (session_id, session['user'], rating, comments))
            conn.commit()
            
            flash("Thank you for your feedback!", "success")
            return redirect(url_for('sit_in_history'))
        
        return render_template('submit_feedback.html', session=session_data)
        
    finally:
        conn.close()

@app.route('/admin/reports/export/sessions/<student_id>')
def export_sessions_csv(student_id):
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    try:
        # Get student info for filename
        student = conn.execute("""
            SELECT lastname, firstname 
            FROM students 
            WHERE idno = ?
        """, (student_id,)).fetchone()
        
        if not student:
            flash("Student not found.", "danger")
            return redirect(url_for('admin_reports'))
        
        # Get student's sit-in sessions
        sessions = conn.execute("""
            SELECT 
                s.idno as "ID Number",
                s.lastname || ', ' || s.firstname || ' ' || COALESCE(s.midname, '') as "Name",
                ls.notes as "Purpose",
                l.room_number || ' - ' || l.building as "Laboratory Room",
                ls.check_in_time as "Check-in Time",
                ls.check_out_time as "Check-out Time",
                DATE(ls.check_in_time) as "Date",
                ROUND(CAST((julianday(ls.check_out_time) - julianday(ls.check_in_time)) * 24 AS REAL), 2) as "Duration (Hours)"
            FROM lab_sessions ls
            JOIN students s ON ls.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE ls.student_id = ? AND ls.status = 'Completed'
            ORDER BY ls.check_in_time DESC
        """, (student_id,)).fetchall()
        
        # Create CSV in memory
        si = StringIO()
        writer = csv.writer(si)
        
        # Write headers
        headers = [
            "ID Number",
            "Name",
            "Purpose",
            "Laboratory Room",
            "Check-in Time",
            "Check-out Time",
            "Date",
            "Duration (Hours)"
        ]
        writer.writerow(headers)
        
        # Write data
        for row in sessions:
            writer.writerow([row[col] for col in row.keys()])
        
        # Create the response
        output = si.getvalue()
        si.close()
        
        # Generate filename
        filename = f"sit_in_sessions_{student['lastname']}_{student['firstname']}_{datetime.now().strftime('%Y-%m-%d')}.csv"
        
        return output, 200, {
            'Content-Type': 'text/csv',
            'Content-Disposition': f'attachment; filename={filename}'
        }
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_reports'))
    finally:
        conn.close()

@app.route('/admin/reports/export/feedback/<student_id>')
def export_feedback_csv(student_id):
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    try:
        # Get student info for filename
        student = conn.execute("""
            SELECT lastname, firstname 
            FROM students 
            WHERE idno = ?
        """, (student_id,)).fetchone()
        
        if not student:
            flash("Student not found.", "danger")
            return redirect(url_for('admin_reports'))
        
        # Get student's feedback
        feedback = conn.execute("""
            SELECT 
                s.idno as "ID Number",
                s.lastname || ', ' || s.firstname || ' ' || COALESCE(s.midname, '') as "Name",
                l.room_number || ' - ' || l.building as "Laboratory Room",
                DATE(ls.check_in_time) as "Date",
                sf.rating as "Rating",
                sf.comments as "Comments"
            FROM session_feedback sf
            JOIN lab_sessions ls ON sf.session_id = ls.session_id
            JOIN students s ON sf.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE sf.student_id = ?
            ORDER BY ls.check_in_time DESC
        """, (student_id,)).fetchall()
        
        # Create CSV in memory
        si = StringIO()
        writer = csv.writer(si)
        
        # Write headers
        headers = [
            "ID Number",
            "Name",
            "Laboratory Room",
            "Date",
            "Rating",
            "Comments"
        ]
        writer.writerow(headers)
        
        # Write data
        for row in feedback:
            writer.writerow([row[col] for col in row.keys()])
        
        # Create the response
        output = si.getvalue()
        si.close()
        
        # Generate filename
        filename = f"feedback_{student['lastname']}_{student['firstname']}_{datetime.now().strftime('%Y-%m-%d')}.csv"
        
        return output, 200, {
            'Content-Type': 'text/csv',
            'Content-Disposition': f'attachment; filename={filename}'
        }
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_reports'))
    finally:
        conn.close()

@app.route('/admin/reports/export/sessions/excel/<student_id>')
def export_sessions_excel(student_id):
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    try:
        # Get student info for filename
        student = conn.execute("""
            SELECT lastname, firstname 
            FROM students 
            WHERE idno = ?
        """, (student_id,)).fetchone()
        
        if not student:
            flash("Student not found.", "danger")
            return redirect(url_for('admin_reports'))
        
        # Get student's sit-in sessions
        sessions = conn.execute("""
            SELECT 
                s.idno as "ID Number",
                s.lastname || ', ' || s.firstname || ' ' || COALESCE(s.midname, '') as "Name",
                ls.notes as "Purpose",
                l.room_number || ' - ' || l.building as "Laboratory Room",
                ls.check_in_time as "Check-in Time",
                ls.check_out_time as "Check-out Time",
                DATE(ls.check_in_time) as "Date",
                ROUND(CAST((julianday(ls.check_out_time) - julianday(ls.check_in_time)) * 24 AS REAL), 2) as "Duration (Hours)"
            FROM lab_sessions ls
            JOIN students s ON ls.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE ls.student_id = ? AND ls.status = 'Completed'
            ORDER BY ls.check_in_time DESC
        """, (student_id,)).fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sit-in Sessions"
        
        # Define headers
        headers = [
            "ID Number",
            "Name",
            "Purpose",
            "Laboratory Room",
            "Check-in Time",
            "Check-out Time",
            "Date",
            "Duration (Hours)"
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row in enumerate(sessions, 2):
            for col_idx, col in enumerate(headers):
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=row[col])
                cell.alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"sit_in_sessions_{student['lastname']}_{student['firstname']}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_reports'))
    finally:
        conn.close()

@app.route('/admin/reports/export/feedback/excel/<student_id>')
def export_feedback_excel(student_id):
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    try:
        # Get student info for filename
        student = conn.execute("""
            SELECT lastname, firstname 
            FROM students 
            WHERE idno = ?
        """, (student_id,)).fetchone()
        
        if not student:
            flash("Student not found.", "danger")
            return redirect(url_for('admin_reports'))
        
        # Get student's feedback
        feedback = conn.execute("""
            SELECT 
                s.idno as "ID Number",
                s.lastname || ', ' || s.firstname || ' ' || COALESCE(s.midname, '') as "Name",
                l.room_number || ' - ' || l.building as "Laboratory Room",
                DATE(ls.check_in_time) as "Date",
                sf.rating as "Rating",
                sf.comments as "Comments"
            FROM session_feedback sf
            JOIN lab_sessions ls ON sf.session_id = ls.session_id
            JOIN students s ON sf.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE sf.student_id = ?
            ORDER BY ls.check_in_time DESC
        """, (student_id,)).fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Feedback"
        
        # Define headers
        headers = [
            "ID Number",
            "Name",
            "Laboratory Room",
            "Date",
            "Rating",
            "Comments"
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row in enumerate(feedback, 2):
            for col_idx, col in enumerate(headers):
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=row[col])
                # Center align all columns except Comments
                if col != "Comments":
                    cell.alignment = Alignment(horizontal='center')
                # For Rating column, convert number to stars
                if col == "Rating":
                    cell.value = "★" * row[col] + "☆" * (5 - row[col])
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"feedback_{student['lastname']}_{student['firstname']}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_reports'))
    finally:
        conn.close()

@app.route('/admin/reports/export/sessions/pdf/<student_id>')
def export_sessions_pdf(student_id):
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    try:
        # Get student info for filename and header
        student = conn.execute("""
            SELECT lastname, firstname 
            FROM students 
            WHERE idno = ?
        """, (student_id,)).fetchone()
        
        if not student:
            flash("Student not found.", "danger")
            return redirect(url_for('admin_reports'))
        
        # Get student's sit-in sessions
        sessions = conn.execute("""
            SELECT 
                s.idno,
                s.lastname,
                s.firstname,
                s.midname,
                ls.notes as purpose,
                l.room_number,
                l.building,
                ls.check_in_time,
                ls.check_out_time,
                DATE(ls.check_in_time) as date,
                ROUND((JULIANDAY(ls.check_out_time) - JULIANDAY(ls.check_in_time)) * 24, 2) as duration
            FROM lab_sessions ls
            JOIN students s ON ls.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE ls.student_id = ? AND ls.status = 'Completed'
            ORDER BY ls.check_in_time DESC
        """, (student_id,)).fetchall()

        # Create PDF buffer
        buffer = BytesIO()
        
        # Create the PDF document - use landscape for better fit
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.darkgreen,
            spaceBefore=10,
            spaceAfter=20,
            alignment=1  # center alignment
        )
        
        # Add the header with logos
        header_data = [
            [
                # UC Logo
                Image(os.path.join('static', 'images', 'uc_logo.png'), width=0.8*inch, height=0.8*inch),
                [
                    Paragraph("University of Cebu - Main Campus", 
                               ParagraphStyle('Header1', parent=styles['Heading2'], alignment=1)),
                    Paragraph("College of Computer Studies", 
                               ParagraphStyle('Header2', parent=styles['Heading3'], alignment=1)),
                    Paragraph("LABORATORY SIT-IN MONITORING SYSTEM REPORT", 
                               ParagraphStyle('Header3', parent=styles['Heading4'], alignment=1, spaceBefore=6))
                ],
                # CCS Logo
                Image(os.path.join('static', 'images', 'ccs_logo.png'), width=0.8*inch, height=0.8*inch)
            ]
        ]
        header_table = Table(header_data, colWidths=[1.2*inch, 8*inch, 1.2*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (1, 0), (1, 0), colors.white),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 20))
        
        # Add title
        title = Paragraph(f"Sit-in Sessions Report - {student['lastname']}, {student['firstname']}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Create paragraph style for table cells to allow word wrapping
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,  # Line spacing
            wordWrap='CJK'  # Enable word wrapping
        )
        
        # Define table data
        data = []
        # Add header row with better column names
        headers = ['ID No.', 'Name', 'Purpose', 'Laboratory', 'Check-in', 'Check-out', 'Date', 'Duration']
        data.append(headers)
        
        # Add session data rows
        for session_row in sessions:
            session_dict = dict(session_row)
            name = f"{session_dict['lastname']}, {session_dict['firstname']} {session_dict['midname'] or ''}"
            laboratory = f"{session_dict['building']} - Room {session_dict['room_number']}"
            
            # Format times for better display
            check_in = format_time(session_dict['check_in_time'])
            check_out = format_time(session_dict['check_out_time'])
            
            # Wrap purpose text in Paragraph for auto-wrapping
            purpose = Paragraph(session_dict['purpose'] or "", cell_style)
            name_para = Paragraph(name, cell_style)
            lab_para = Paragraph(laboratory, cell_style)
            
            data.append([
                session_dict['idno'],
                name_para,
                purpose,
                lab_para,
                check_in,
                check_out,
                session_dict['date'],
                f"{session_dict['duration']} hrs"
            ])
        
        # Optimize column widths
        col_widths = [
            0.8*inch,      # ID No
            1.5*inch,      # Name
            2.0*inch,      # Purpose (wider for text wrapping)
            1.5*inch,      # Laboratory
            0.9*inch,      # Check-in
            0.9*inch,      # Check-out
            0.9*inch,      # Date
            0.8*inch       # Duration
        ]
        
        # Create the table with the data
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Add style to the table
        style = TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows styling - alternating colors
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            
            # Column alignments
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # ID column
            ('ALIGN', (4, 1), (7, -1), 'CENTER'),  # Time, date, duration columns
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Grid styling
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            
            # Cell padding
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ])
        
        # Add alternating row colors
        for i in range(1, len(data)):
            if i % 2 == 0:
                style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)
        
        table.setStyle(style)
        
        # Add the table to the elements
        elements.append(table)
        
        # Add footer with page numbers
        def add_page_number(canvas, doc):
            page_num = canvas.getPageNumber()
            text = f"Page {page_num}"
            canvas.setFont("Helvetica", 8)
            canvas.drawRightString(doc.width + doc.rightMargin - 10, doc.bottomMargin - 20, text)
            canvas.drawString(doc.leftMargin, doc.bottomMargin - 20, 
                              f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Build the PDF with page numbers
        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
        
        # Get the value of the BytesIO buffer
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Create the response
        response = make_response(pdf_data)
        response.headers['Content-Disposition'] = f'attachment; filename=sit_in_sessions_{student["lastname"]}_{student["firstname"]}_{datetime.now().strftime("%Y-%m-%d")}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        
        return response
        
    except Exception as e:
        print(f"Error exporting sessions PDF: {str(e)}")
        flash('An error occurred while exporting the data.', 'error')
        return redirect(url_for('admin_reports'))

@app.route('/admin/reports/export/feedback/pdf/<student_id>')
def export_feedback_pdf(student_id):
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        
        # Get student info for filename and header
        student = conn.execute("""
            SELECT lastname, firstname, course, year_level
            FROM students 
            WHERE idno = ?
        """, (student_id,)).fetchone()
        
        if not student:
            flash("Student not found.", "danger")
            return redirect(url_for('admin_reports'))
        
        # Get student's feedback submissions
        feedback = conn.execute("""
            SELECT 
                s.idno,
                s.lastname,
                s.firstname,
                s.midname,
                l.building,
                l.room_number,
                DATE(ls.check_in_time) as date,
                sf.rating,
                sf.comments
            FROM session_feedback sf
            JOIN lab_sessions ls ON sf.session_id = ls.session_id
            JOIN students s ON sf.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE sf.student_id = ?
            ORDER BY ls.check_in_time DESC
        """, (student_id,)).fetchall()

        # Create PDF buffer
        buffer = BytesIO()
        
        # Create the PDF document - use landscape for better fit
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.darkgreen,
            spaceBefore=10,
            spaceAfter=20,
            alignment=1  # center alignment
        )
        
        # Define cell style for wrapping text
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10
        )
        
        # Add the header with logos
        header_data = [
            [
                # UC Logo
                Image(os.path.join('static', 'images', 'uc_logo.png'), width=0.7*inch, height=0.7*inch),
                [
                    Paragraph("University of Cebu - Main Campus", 
                               ParagraphStyle('Header1', parent=styles['Heading2'], alignment=1, fontSize=12)),
                    Paragraph("College of Computer Studies", 
                               ParagraphStyle('Header2', parent=styles['Heading3'], alignment=1, fontSize=10)),
                    Paragraph("LABORATORY SIT-IN MONITORING SYSTEM REPORT", 
                               ParagraphStyle('Header3', parent=styles['Heading4'], alignment=1, fontSize=9, spaceBefore=4))
                ],
                # CCS Logo
                Image(os.path.join('static', 'images', 'ccs_logo.png'), width=0.7*inch, height=0.7*inch)
            ]
        ]
        header_table = Table(header_data, colWidths=[1*inch, 8*inch, 1*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (1, 0), (1, 0), colors.white),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 20))
        
        # Add title and student info
        title = Paragraph(f"Feedback Report - {student['lastname']}, {student['firstname']}", title_style)
        elements.append(title)
        
        # Add student info table
        student_info = [
            ["ID Number:", student_id, "Course:", student['course']],
            ["Name:", f"{student['lastname']}, {student['firstname']}", "Year Level:", student['year_level']],
        ]
        student_table = Table(student_info, colWidths=[1*inch, 3*inch, 1*inch, 3*inch])
        student_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(student_table)
        elements.append(Spacer(1, 12))
        
        # Define table data
        data = []
        # Add header row
        headers = ['Laboratory', 'Date', 'Rating', 'Comments']
        data.append(headers)
        
        # Add feedback data rows
        for feedback_row in feedback:
            feedback_dict = dict(feedback_row)
            laboratory = f"{feedback_dict['building']} - Room {feedback_dict['room_number']}"
            stars = "★" * feedback_dict['rating'] + "☆" * (5 - feedback_dict['rating'])
            
            # Wrap text in Paragraph for auto-wrapping
            lab_para = Paragraph(laboratory, cell_style)
            comments_para = Paragraph(feedback_dict['comments'] or "", cell_style)
            
            data.append([
                lab_para,
                feedback_dict['date'],
                stars,
                comments_para
            ])
        
        # Optimize column widths
        col_widths = [
            2.0*inch,      # Laboratory
            1.5*inch,      # Date
            1.0*inch,      # Rating
            5.5*inch       # Comments
        ]
        
        # Create the table with the data
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Add style to the table
        style = TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows styling - alternating colors
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            
            # Column alignments
            ('ALIGN', (1, 1), (2, -1), 'CENTER'),  # Date and Rating columns
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Grid styling
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            
            # Cell padding
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ])
        
        # Add alternating row colors
        for i in range(1, len(data)):
            if i % 2 == 0:
                style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)
        
        table.setStyle(style)
        
        # Add the table to the elements
        elements.append(table)
        
        # Add footer with page numbers
        def add_page_number(canvas, doc):
            page_num = canvas.getPageNumber()
            text = f"Page {page_num}"
            canvas.setFont("Helvetica", 8)
            canvas.drawRightString(doc.width + doc.rightMargin - 10, doc.bottomMargin - 20, text)
            canvas.drawString(doc.leftMargin, doc.bottomMargin - 20, 
                             f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Build the PDF with page numbers
        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
        
        # Get the value of the BytesIO buffer
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Create the response
        response = make_response(pdf_data)
        response.headers['Content-Disposition'] = f'attachment; filename=feedback_{student["lastname"]}_{student["firstname"]}_{datetime.now().strftime("%Y-%m-%d")}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        
        return response
        
    except Exception as e:
        print(f"Error exporting feedback PDF: {str(e)}")
        flash('An error occurred while exporting the data.', 'error')
        return redirect(url_for('admin_reports'))

@app.route('/admin/reports/all')
def admin_all_reports():
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    try:
        # Fetch all students
        students = conn.execute("""
            SELECT idno, lastname, firstname, midname, course, year_level, email_address
            FROM students
            ORDER BY lastname, firstname
        """).fetchall()
        
        # Get aggregate data for all students
        # 1. Total completed sessions
        total_sessions = conn.execute("""
            SELECT COUNT(*) as count
            FROM lab_sessions
            WHERE status = 'Completed'
        """).fetchone()['count']
        
        # 2. Total feedback submissions
        total_feedback = conn.execute("""
            SELECT COUNT(*) as count
            FROM session_feedback
        """).fetchone()['count']
        
        # 3. Laboratory usage statistics
        lab_stats = conn.execute("""
            SELECT 
                l.building || ' - Room ' || l.room_number as lab,
                COUNT(*) as count
            FROM lab_sessions ls
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE ls.status = 'Completed'
            GROUP BY l.building, l.room_number
            ORDER BY count DESC
        """).fetchall()
        
        # 4. Purpose statistics
        purpose_stats = conn.execute("""
            SELECT 
                notes as purpose,
                COUNT(*) as count
            FROM lab_sessions
            WHERE status = 'Completed'
            GROUP BY notes
            ORDER BY count DESC
        """).fetchall()
        
        # 5. Recent sessions (limit to 20)
        recent_sessions = conn.execute("""
            SELECT 
                s.idno,
                s.lastname || ', ' || s.firstname || ' ' || COALESCE(s.midname, '') as name,
                ls.notes as purpose,
                l.room_number || ' - ' || l.building as lab,
                ls.check_in_time,
                ls.check_out_time,
                DATE(ls.check_in_time) as date,
                ROUND(CAST((julianday(ls.check_out_time) - julianday(ls.check_in_time)) * 24 AS REAL), 2) as duration
            FROM lab_sessions ls
            JOIN students s ON ls.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE ls.status = 'Completed'
            ORDER BY ls.check_in_time DESC
            LIMIT 20
        """).fetchall()
        
        # 6. Recent feedback (limit to 20)
        recent_feedback = conn.execute("""
            SELECT 
                s.idno,
                s.lastname || ', ' || s.firstname || ' ' || COALESCE(s.midname, '') as name,
                l.room_number || ' - ' || l.building as lab,
                DATE(ls.check_in_time) as date,
                sf.rating,
                sf.comments
            FROM session_feedback sf
            JOIN lab_sessions ls ON sf.session_id = ls.session_id
            JOIN students s ON sf.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            ORDER BY ls.check_in_time DESC
            LIMIT 20
        """).fetchall()
        
        return render_template(
            'admin_all_reports.html',
            students=students,
            total_sessions=total_sessions,
            total_feedback=total_feedback,
            lab_stats=lab_stats,
            purpose_stats=purpose_stats,
            recent_sessions=recent_sessions,
            recent_feedback=recent_feedback
        )
    
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_reports'))
    finally:
        conn.close()

@app.route('/admin/reports/export/all/sessions/csv')
def export_all_sessions_csv():
    # Check if user is logged in and is admin
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = None
    try:
        conn = get_db_connection()
        
        # Get all sessions
        sessions = conn.execute("""
            SELECT 
                s.idno,
                s.lastname,
                s.firstname,
                s.midname,
                ls.notes as purpose,
                l.building,
                l.room_number,
                ls.check_in_time,
                ls.check_out_time,
                DATE(ls.check_in_time) as date,
                ROUND((JULIANDAY(ls.check_out_time) - JULIANDAY(ls.check_in_time)) * 24, 2) as duration
            FROM lab_sessions ls
            JOIN students s ON ls.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE ls.check_out_time IS NOT NULL
            ORDER BY ls.check_in_time DESC
        """).fetchall()
        
        # Create CSV
        csv_data = io.StringIO()
        csv_writer = csv.writer(csv_data)
        
        # Write headers
        headers = ['ID Number', 'Last Name', 'First Name', 'Middle Name', 'Purpose', 'Building', 'Room Number', 'Check-in Time', 'Check-out Time', 'Date', 'Duration (Hours)']
        csv_writer.writerow(headers)
        
        # Write data
        for session_row in sessions:
            session_dict = dict(session_row)
            csv_writer.writerow([
                session_dict['idno'],
                session_dict['lastname'],
                session_dict['firstname'],
                session_dict['midname'] or "",
                session_dict['purpose'] or "",
                session_dict['building'],
                session_dict['room_number'],
                session_dict['check_in_time'],
                session_dict['check_out_time'],
                session_dict['date'],
                session_dict['duration']
            ])
        
        # Prepare response
        response = make_response(csv_data.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=all_sit_in_sessions_{datetime.now().strftime("%Y-%m-%d")}.csv'
        response.headers['Content-Type'] = 'text/csv'
        
        return response
    
    except Exception as e:
        print(f"Error exporting all sessions CSV: {str(e)}")
        flash(f'An error occurred while exporting the data: {str(e)}', 'error')
        return redirect(url_for('admin_all_reports'))
    finally:
        if conn:
            conn.close()

@app.route('/admin/reports/export/all/feedback/csv')
def export_all_feedback_csv():
    # Check if user is logged in and is admin
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = None
    try:
        conn = get_db_connection()
        
        # Get all feedback
        feedback = conn.execute("""
            SELECT 
                s.idno,
                s.lastname,
                s.firstname,
                s.midname,
                l.building,
                l.room_number,
                DATE(ls.check_in_time) as date,
                sf.rating,
                sf.comments
            FROM session_feedback sf
            JOIN lab_sessions ls ON sf.session_id = ls.session_id
            JOIN students s ON sf.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            ORDER BY ls.check_in_time DESC
        """).fetchall()
        
        # Create CSV
        csv_data = io.StringIO()
        csv_writer = csv.writer(csv_data)
        
        # Write headers
        headers = ['ID Number', 'Last Name', 'First Name', 'Middle Name', 'Building', 'Room Number', 'Date', 'Rating', 'Comments']
        csv_writer.writerow(headers)
        
        # Write data
        for feedback_row in feedback:
            feedback_dict = dict(feedback_row)
            csv_writer.writerow([
                feedback_dict['idno'],
                feedback_dict['lastname'],
                feedback_dict['firstname'],
                feedback_dict['midname'] or "",
                feedback_dict['building'],
                feedback_dict['room_number'],
                feedback_dict['date'],
                feedback_dict['rating'],
                feedback_dict['comments'] or ""
            ])
        
        # Prepare response
        response = make_response(csv_data.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=all_feedback_{datetime.now().strftime("%Y-%m-%d")}.csv'
        response.headers['Content-Type'] = 'text/csv'
        
        return response
    
    except Exception as e:
        print(f"Error exporting all feedback CSV: {str(e)}")
        flash(f'An error occurred while exporting the data: {str(e)}', 'error')
        return redirect(url_for('admin_all_reports'))
    finally:
        if conn:
            conn.close()

@app.route('/admin/reports/export/all/sessions/excel')
def export_all_sessions_excel():
    # Check if user is logged in and is admin
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = None
    try:
        conn = get_db_connection()
        
        # Get all sessions
        sessions = conn.execute("""
            SELECT 
                s.idno,
                s.lastname,
                s.firstname,
                s.midname,
                ls.notes as purpose,
                l.building,
                l.room_number,
                ls.check_in_time,
                ls.check_out_time,
                DATE(ls.check_in_time) as date,
                ROUND((JULIANDAY(ls.check_out_time) - JULIANDAY(ls.check_in_time)) * 24, 2) as duration
            FROM lab_sessions ls
            JOIN students s ON ls.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE ls.check_out_time IS NOT NULL
            ORDER BY ls.check_in_time DESC
        """).fetchall()
        
        # Create a BytesIO object
        output = BytesIO()
        
        # Use pandas to create an Excel file
        import pandas as pd
        
        # Convert data to pandas DataFrame
        data = []
        for session_row in sessions:
            session_dict = dict(session_row)
            data.append({
                'ID Number': session_dict['idno'],
                'Last Name': session_dict['lastname'],
                'First Name': session_dict['firstname'],
                'Middle Name': session_dict['midname'] or "",
                'Purpose': session_dict['purpose'] or "",
                'Building': session_dict['building'],
                'Room Number': session_dict['room_number'],
                'Check-in Time': session_dict['check_in_time'],
                'Check-out Time': session_dict['check_out_time'],
                'Date': session_dict['date'],
                'Duration (Hours)': session_dict['duration']
            })
        
        df = pd.DataFrame(data)
        
        # Write DataFrame to Excel
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='All Sessions', index=False)
            
            # Access the workbook and the worksheet
            workbook = writer.book
            worksheet = writer.sheets['All Sessions']
            
            # Add a header format
            header_format = workbook.add_format({
                'bold': True,
                'fg_color': '#4CAF50',
                'font_color': 'white',
                'border': 1
            })
            
            # Apply header format to the header row
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                
            # Auto-adjust column widths
            for i, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).str.len().max(), len(col)) + 2
                worksheet.set_column(i, i, max_len)
        
        # Prepare response
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=all_sit_in_sessions_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
        
    except Exception as e:
        print(f"Error exporting all sessions Excel: {str(e)}")
        flash(f'An error occurred while exporting the data: {str(e)}', 'error')
        return redirect(url_for('admin_all_reports'))
    finally:
        if conn:
            conn.close()

@app.route('/admin/reports/export/all/feedback/excel')
def export_all_feedback_excel():
    # Check if user is logged in and is admin
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = None
    try:
        conn = get_db_connection()
        
        # Get all feedback
        feedback = conn.execute("""
            SELECT 
                s.idno,
                s.lastname,
                s.firstname,
                s.midname,
                l.building,
                l.room_number,
                DATE(ls.check_in_time) as date,
                sf.rating,
                sf.comments
            FROM session_feedback sf
            JOIN lab_sessions ls ON sf.session_id = ls.session_id
            JOIN students s ON sf.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            ORDER BY ls.check_in_time DESC
        """).fetchall()
        
        # Create a BytesIO object
        output = BytesIO()
        
        # Use pandas to create an Excel file
        import pandas as pd
        
        # Convert data to pandas DataFrame
        data = []
        for feedback_row in feedback:
            feedback_dict = dict(feedback_row)
            # Convert rating to stars
            stars = "★" * feedback_dict['rating'] + "☆" * (5 - feedback_dict['rating'])
            data.append({
                'ID Number': feedback_dict['idno'],
                'Last Name': feedback_dict['lastname'],
                'First Name': feedback_dict['firstname'],
                'Middle Name': feedback_dict['midname'] or "",
                'Building': feedback_dict['building'],
                'Room Number': feedback_dict['room_number'],
                'Date': feedback_dict['date'],
                'Rating': stars,
                'Comments': feedback_dict['comments'] or ""
            })
        
        df = pd.DataFrame(data)
        
        # Write DataFrame to Excel
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='All Feedback', index=False)
            
            # Access the workbook and the worksheet
            workbook = writer.book
            worksheet = writer.sheets['All Feedback']
            
            # Add a header format
            header_format = workbook.add_format({
                'bold': True,
                'fg_color': '#4CAF50',
                'font_color': 'white',
                'border': 1
            })
            
            # Apply header format to the header row
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                
            # Auto-adjust column widths
            for i, col in enumerate(df.columns):
                if col == 'Comments':
                    worksheet.set_column(i, i, 40)  # Make comments column wider
                else:
                    max_len = max(df[col].astype(str).str.len().max(), len(col)) + 2
                    worksheet.set_column(i, i, max_len)
        
        # Prepare response
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=all_feedback_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
        
    except Exception as e:
        print(f"Error exporting all feedback Excel: {str(e)}")
        flash(f'An error occurred while exporting the data: {str(e)}', 'error')
        return redirect(url_for('admin_all_reports'))
    finally:
        if conn:
            conn.close()

@app.route('/admin/reports/export/all/sessions/pdf')
def export_all_sessions_pdf():
    # Check if user is logged in and is admin
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = None
    try:
        conn = get_db_connection()
        
        # Get all sessions
        sessions = conn.execute("""
            SELECT 
                s.idno,
                s.lastname,
                s.firstname,
                s.midname,
                ls.notes as purpose,
                l.building,
                l.room_number,
                ls.check_in_time,
                ls.check_out_time,
                DATE(ls.check_in_time) as date,
                ROUND((JULIANDAY(ls.check_out_time) - JULIANDAY(ls.check_in_time)) * 24, 2) as duration
            FROM lab_sessions ls
            JOIN students s ON ls.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            WHERE ls.check_out_time IS NOT NULL
            ORDER BY ls.check_in_time DESC
        """).fetchall()
        
        # Create PDF buffer
        buffer = BytesIO()
        
        # Create the PDF document - use landscape for better fit
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.darkgreen,
            spaceBefore=10,
            spaceAfter=20,
            alignment=1  # center alignment
        )
        
        # Add the header with logos
        header_data = [
            [
                # UC Logo
                Image(os.path.join('static', 'images', 'uc_logo.png'), width=0.8*inch, height=0.8*inch),
                [
                    Paragraph("University of Cebu - Main Campus", 
                               ParagraphStyle('Header1', parent=styles['Heading2'], alignment=1)),
                    Paragraph("College of Computer Studies", 
                               ParagraphStyle('Header2', parent=styles['Heading3'], alignment=1)),
                    Paragraph("LABORATORY SIT-IN MONITORING SYSTEM REPORT", 
                               ParagraphStyle('Header3', parent=styles['Heading4'], alignment=1, spaceBefore=6))
                ],
                # CCS Logo
                Image(os.path.join('static', 'images', 'ccs_logo.png'), width=0.8*inch, height=0.8*inch)
            ]
        ]
        header_table = Table(header_data, colWidths=[1.2*inch, 8*inch, 1.2*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (1, 0), (1, 0), colors.white),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 20))
        
        # Add title
        title = Paragraph(f"All Sit-in Sessions - {datetime.now().strftime('%Y-%m-%d')}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Create paragraph style for table cells to allow word wrapping
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,  # Line spacing
            wordWrap='CJK'  # Enable word wrapping
        )
        
        # Define table data
        data = []
        # Add header row
        headers = ['ID No.', 'Name', 'Purpose', 'Laboratory', 'Check-in', 'Check-out', 'Date', 'Duration']
        data.append(headers)
        
        # Add session data rows
        for session_row in sessions:
            session_dict = dict(session_row)
            name = f"{session_dict['lastname']}, {session_dict['firstname']} {session_dict['midname'] or ''}"
            laboratory = f"{session_dict['building']} - Room {session_dict['room_number']}"
            
            # Format times for better display
            check_in = format_time(session_dict['check_in_time'])
            check_out = format_time(session_dict['check_out_time'])
            
            # Wrap text in Paragraph for auto-wrapping
            purpose = Paragraph(session_dict['purpose'] or "", cell_style)
            name_para = Paragraph(name, cell_style)
            lab_para = Paragraph(laboratory, cell_style)
            
            data.append([
                session_dict['idno'],
                name_para,
                purpose,
                lab_para,
                check_in,
                check_out,
                session_dict['date'],
                f"{session_dict['duration']} hrs"
            ])
        
        # Optimize column widths
        col_widths = [
            0.8*inch,      # ID No
            1.5*inch,      # Name
            2.0*inch,      # Purpose (wider for text wrapping)
            1.5*inch,      # Laboratory
            0.9*inch,      # Check-in
            0.9*inch,      # Check-out
            0.9*inch,      # Date
            0.8*inch       # Duration
        ]
        
        # Create the table with the data
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Add style to the table
        style = TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows styling - alternating colors
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            
            # Column alignments
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # ID column
            ('ALIGN', (4, 1), (7, -1), 'CENTER'),  # Time, date, duration columns
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Grid styling
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            
            # Cell padding
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ])
        
        # Add alternating row colors
        for i in range(1, len(data)):
            if i % 2 == 0:
                style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)
        
        table.setStyle(style)
        
        # Add the table to the elements
        elements.append(table)
        
        # Build the PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Create the response
        response = make_response(pdf_data)
        response.headers['Content-Disposition'] = f'attachment; filename=all_sit_in_sessions_{datetime.now().strftime("%Y-%m-%d")}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        
        return response
        
    except Exception as e:
        print(f"Error exporting all sessions PDF: {str(e)}")
        flash(f'An error occurred while exporting the data: {str(e)}', 'error')
        return redirect(url_for('admin_all_reports'))
    finally:
        if conn:
            conn.close()

@app.route('/admin/reports/export/all/feedback/pdf')
def export_all_feedback_pdf():
    # Check if user is logged in and is admin
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = None
    try:
        conn = get_db_connection()
        
        # Get all feedback
        feedback = conn.execute("""
            SELECT 
                s.idno,
                s.lastname,
                s.firstname,
                s.midname,
                s.course,
                s.year_level,
                l.building,
                l.room_number,
                DATE(ls.check_in_time) as date,
                sf.rating,
                sf.comments
            FROM session_feedback sf
            JOIN lab_sessions ls ON sf.session_id = ls.session_id
            JOIN students s ON sf.student_id = s.idno
            JOIN laboratories l ON ls.lab_id = l.lab_id
            ORDER BY ls.check_in_time DESC
        """).fetchall()
        
        # Create PDF buffer
        buffer = BytesIO()
        
        # Create the PDF document - use landscape for better fit
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.darkgreen,
            spaceBefore=10,
            spaceAfter=20,
            alignment=1  # center alignment
        )
        
        # Define cell style for wrapping text
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10
        )
        
        # Add the header with logos
        header_data = [
            [
                # UC Logo
                Image(os.path.join('static', 'images', 'uc_logo.png'), width=0.7*inch, height=0.7*inch),
                [
                    Paragraph("University of Cebu - Main Campus", 
                               ParagraphStyle('Header1', parent=styles['Heading2'], alignment=1, fontSize=12)),
                    Paragraph("College of Computer Studies", 
                               ParagraphStyle('Header2', parent=styles['Heading3'], alignment=1, fontSize=10)),
                    Paragraph("LABORATORY SIT-IN MONITORING SYSTEM REPORT", 
                               ParagraphStyle('Header3', parent=styles['Heading4'], alignment=1, fontSize=9, spaceBefore=4))
                ],
                # CCS Logo
                Image(os.path.join('static', 'images', 'ccs_logo.png'), width=0.7*inch, height=0.7*inch)
            ]
        ]
        header_table = Table(header_data, colWidths=[1*inch, 8*inch, 1*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (1, 0), (1, 0), colors.white),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 20))
        
        # Add title
        title = Paragraph(f"All Student Feedback - {datetime.now().strftime('%Y-%m-%d')}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Define table data
        data = []
        # Add header row
        headers = ['ID No.', 'Name', 'Course', 'Year', 'Laboratory', 'Date', 'Rating', 'Comments']
        data.append(headers)
        
        # Add feedback data rows
        for feedback_row in feedback:
            feedback_dict = dict(feedback_row)
            name = f"{feedback_dict['lastname']}, {feedback_dict['firstname']} {feedback_dict['midname'] or ''}"
            laboratory = f"{feedback_dict['building']} - Room {feedback_dict['room_number']}"
            stars = "★" * feedback_dict['rating'] + "☆" * (5 - feedback_dict['rating'])
            
            # Wrap text in Paragraph for auto-wrapping
            name_para = Paragraph(name, cell_style)
            lab_para = Paragraph(laboratory, cell_style)
            comments_para = Paragraph(feedback_dict['comments'] or "", cell_style)
            
            data.append([
                feedback_dict['idno'],
                name_para,
                feedback_dict['course'],
                feedback_dict['year_level'],
                lab_para,
                feedback_dict['date'],
                stars,
                comments_para
            ])
        
        # Optimize column widths
        col_widths = [
            0.8*inch,      # ID No
            1.5*inch,      # Name
            0.9*inch,      # Course
            0.5*inch,      # Year
            1.3*inch,      # Laboratory
            0.8*inch,      # Date
            0.7*inch,      # Rating
            3.5*inch       # Comments
        ]
        
        # Create the table with the data
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Add style to the table
        style = TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows styling - alternating colors
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            
            # Column alignments
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # ID column
            ('ALIGN', (2, 1), (3, -1), 'CENTER'),  # Course and Year columns
            ('ALIGN', (5, 1), (6, -1), 'CENTER'),  # Date and Rating columns
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Grid styling
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            
            # Cell padding
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ])
        table.setStyle(style)
        
        # Add the table to the elements
        elements.append(table)
        
        # Build the PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Create the response
        response = make_response(pdf_data)
        response.headers['Content-Disposition'] = f'attachment; filename=all_feedback_{datetime.now().strftime("%Y-%m-%d")}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        
        return response
        
    except Exception as e:
        print(f"Error exporting all feedback PDF: {str(e)}")
        flash(f'An error occurred while exporting the data: {str(e)}', 'error')
        return redirect(url_for('admin_all_reports'))
    finally:
        if conn:
            conn.close()

@app.route('/admin/edit-student/<student_id>', methods=['GET', 'POST'])
def admin_edit_student(student_id):
    # Check if user is logged in and is admin
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    
    if request.method == 'POST':
        # Retrieve form data
        idno = request.form.get('idno')
        lastname = request.form.get('lastname')
        firstname = request.form.get('firstname')
        midname = request.form.get('midname')
        course = request.form.get('course')
        year_level = request.form.get('year_level')
        email_address = request.form.get('email_address')
        delete_image = request.form.get('delete_image') == 'yes'  # Check if delete checkbox is checked

        # Verify that the student ID matches the URL parameter
        if idno != student_id:
            flash("Student ID mismatch. Operation aborted.", "danger")
            return redirect(url_for('admin_dashboard'))

        # Check for email conflicts
        existing_email = conn.execute(
            "SELECT * FROM students WHERE email_address = ? AND idno != ?",
            (email_address, idno)
        ).fetchone()
        if existing_email:
            flash("Email is already used by another account.", "danger")
            conn.close()
            return redirect(url_for('admin_edit_student', student_id=student_id))
        
        # Get current student info to access current image path
        current_student = conn.execute("SELECT image_path FROM students WHERE idno = ?", (idno,)).fetchone()
        image_path = current_student['image_path']  # Keep current path by default
        
        # Handle profile image upload or deletion
        profile_image = request.files.get('profile_image')
        
        if delete_image and image_path:
            # Delete the physical file if it exists
            if image_path:
                file_path = os.path.join('static', image_path)
                if os.path.exists(file_path):
                    os.remove(file_path)
            # Set image_path to None in the database
            image_path = None
            
        elif profile_image and profile_image.filename:
            # Delete old image file if exists
            if image_path:
                old_file_path = os.path.join('static', image_path)
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)
                    
            # Process new image
            filename = secure_filename(profile_image.filename)
            upload_folder = os.path.join('static', 'uploads')
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)
            save_path = os.path.join(upload_folder, filename)
            profile_image.save(save_path)
            image_path = os.path.join('uploads', filename).replace(os.sep, '/')

        # Update student record including image path
        conn.execute("""
            UPDATE students
            SET lastname = ?,
                firstname = ?,
                midname = ?,
                course = ?,
                year_level = ?,
                email_address = ?,
                image_path = ?
            WHERE idno = ?
        """, (lastname, firstname, midname, course, year_level, email_address, image_path, idno))
        
        conn.commit()
        conn.close()
        flash("Student updated successfully!", "success")
        return redirect(url_for('admin_dashboard'))
    
    else:
        # For GET requests, retrieve the specified student info
        student = conn.execute("SELECT * FROM students WHERE idno = ?", (student_id,)).fetchone()
        conn.close()
        if not student:
            flash("Student not found.", "danger")
            return redirect(url_for('admin_dashboard'))
        return render_template('admin_edit_student.html', student=student)

@app.route('/admin/feedback')
def admin_feedback():
    # Check if user is logged in and is admin
    if 'user' not in session or session.get('is_admin') != True:
        flash("Please log in as an administrator.", "warning")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    
    # Get all feedback with student and lab details
    feedback = conn.execute("""
        SELECT sf.*, s.firstname, s.lastname, s.course, s.year_level, 
               l.building, l.room_number, ls.check_in_time,
               DATE(ls.check_in_time) as date
        FROM session_feedback sf
        JOIN students s ON sf.student_id = s.idno
        JOIN lab_sessions ls ON sf.session_id = ls.session_id
        JOIN laboratories l ON ls.lab_id = l.lab_id
        ORDER BY ls.check_in_time DESC
    """).fetchall()
    
    # Calculate statistics
    total_feedback = len(feedback)
    
    # Calculate average rating
    if total_feedback > 0:
        avg_rating = conn.execute("""
            SELECT AVG(rating) as avg_rating FROM session_feedback
        """).fetchone()['avg_rating']
    else:
        avg_rating = 0
    
    # Get rating distribution
    rating_distribution = []
    for rating in range(1, 6):
        count = conn.execute("""
            SELECT COUNT(*) as count FROM session_feedback
            WHERE rating = ?
        """, (rating,)).fetchone()['count']
        rating_distribution.append({'rating': rating, 'count': count})
    
    conn.close()
    
    return render_template(
        'admin_feedback.html',
        feedback=feedback,
        total_feedback=total_feedback,
        avg_rating=avg_rating,
        rating_distribution=rating_distribution
    )

@app.route('/admin/leaderboard')
def admin_leaderboard():
    # Check if user is logged in and is admin
    if 'user' not in session or session.get('is_admin') != True:
        flash("Please log in as an administrator.", "warning")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    
    # Get top 10 most active students (by completed session count)
    most_active_students = conn.execute("""
        SELECT 
            s.idno, s.firstname, s.lastname, s.course, s.year_level, s.image_path,
            COUNT(ls.session_id) as session_count
        FROM students s
        JOIN lab_sessions ls ON s.idno = ls.student_id
        WHERE ls.status = 'Completed'
        GROUP BY s.idno
        ORDER BY session_count DESC
        LIMIT 10
    """).fetchall()
    
    # Get top 10 students with the most behavior points
    top_performers = conn.execute("""
        SELECT 
            s.idno, s.firstname, s.lastname, s.course, s.year_level, s.image_path,
            COALESCE(SUM(ls.behavior_points), 0) as behavior_points
        FROM students s
        LEFT JOIN lab_sessions ls ON s.idno = ls.student_id AND ls.behavior_points > 0
        GROUP BY s.idno
        HAVING behavior_points > 0
        ORDER BY behavior_points DESC
        LIMIT 10
    """).fetchall()
    
    # Calculate free sessions for each top performer
    top_performers_list = []
    for student in top_performers:
        student_dict = dict(student)
        student_dict['free_sessions'] = student_dict['behavior_points'] // 3
        top_performers_list.append(student_dict)
    
    conn.close()
    
    return render_template(
        'admin_leaderboard.html',
        most_active_students=most_active_students,
        top_performers=top_performers_list
    )

@app.route('/admin/award_behavior_point/<int:session_id>')
def award_behavior_point(session_id):
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    conn = None
    try:
        conn = get_db_connection()
        
        # Check if the session exists
        session_data = conn.execute("SELECT * FROM lab_sessions WHERE session_id = ?", (session_id,)).fetchone()
        
        if not session_data:
            flash("Session not found.", "danger")
            return redirect(url_for('admin_sit_in_records'))
        
        # Check if the student already has a behavior point for this session
        existing_point = conn.execute(
            "SELECT behavior_points FROM lab_sessions WHERE session_id = ?", 
            (session_id,)
        ).fetchone()
        
        if existing_point and existing_point['behavior_points']:
            flash("This student has already been awarded a behavior point for this session.", "warning")
        else:
            # Award the behavior point
            conn.execute(
                "UPDATE lab_sessions SET behavior_points = 1 WHERE session_id = ?",
                (session_id,)
            )
            conn.commit()
            
            # Get student info and calculate total points
            student_info = conn.execute("""
                SELECT s.idno, s.firstname, s.lastname
                FROM lab_sessions ls
                JOIN students s ON ls.student_id = s.idno
                WHERE ls.session_id = ?
            """, (session_id,)).fetchone()
            
            student_name = f"{student_info['firstname']} {student_info['lastname']}" if student_info else "Student"
            
            # Calculate total behavior points for the student
            if student_info:
                total_points = conn.execute("""
                    SELECT COALESCE(SUM(behavior_points), 0) as total
                    FROM lab_sessions
                    WHERE student_id = ? AND behavior_points > 0
                """, (student_info['idno'],)).fetchone()['total']
                
                # Check if student has reached a multiple of 3 points
                if total_points > 0 and total_points % 3 == 0:
                    free_sessions = total_points // 3
                    flash(f"Behavior point awarded to {student_name} successfully! They now have {total_points} points, earning them {free_sessions} free session(s)!", "success")
                else:
                    points_to_next_free = 3 - (total_points % 3)
                    flash(f"Behavior point awarded to {student_name} successfully! They now have {total_points} points. {points_to_next_free} more points needed for another free session.", "success")
            else:
                flash(f"Behavior point awarded successfully!", "success")
        
        # Determine which page to redirect to based on the session status
        if session_data['status'] == 'Active':
            return redirect(url_for('admin_current_sit_in'))
        else:
            return redirect(url_for('admin_sit_in_records'))
            
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_dashboard'))
    finally:
        if conn:
            conn.close()

@app.route('/admin/reset_sessions')
def admin_reset_sessions():
    """Reset all students' remaining sessions to default values based on their course."""
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        
        # Delete all completed lab sessions
        conn.execute("DELETE FROM lab_sessions WHERE status = 'Completed'")
        
        # Clear all lab reservations
        conn.execute("DELETE FROM lab_reservations")
        
        # Reset free_sessions_used for all students
        conn.execute("UPDATE students SET free_sessions_used = 0")
        
        # Log the action
        conn.execute("""
            INSERT INTO admin_logs (admin_id, action, timestamp)
            VALUES (?, 'Reset all student sessions and cleared all reservations', datetime('now'))
        """, (session['user'],))
        
        conn.commit()
        conn.close()
        
        flash("All student sessions have been reset and all reservations have been cleared.", "success")
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
    
    return redirect(url_for('admin_system_management'))

@app.route('/admin/reset_points')
def admin_reset_points():
    """Reset all behavior points for all students."""
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        
        # Set all behavior points to 0
        conn.execute("UPDATE lab_sessions SET behavior_points = 0 WHERE behavior_points > 0")
        
        # Log the action
        conn.execute("""
            INSERT INTO admin_logs (admin_id, action, timestamp)
            VALUES (?, 'Reset all behavior points', datetime('now'))
        """, (session['user'],))
        
        conn.commit()
        conn.close()
        
        flash("All student behavior points have been reset.", "success")
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
    
    return redirect(url_for('admin_system_management'))

@app.route('/admin/reset_all')
def admin_reset_all():
    """Reset both sessions and behavior points for all students."""
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        
        # Delete all completed lab sessions (automatically removes their behavior points)
        conn.execute("DELETE FROM lab_sessions WHERE status = 'Completed'")
        
        # Also reset behavior points for any active sessions
        conn.execute("UPDATE lab_sessions SET behavior_points = 0 WHERE behavior_points > 0")
        
        # Clear all lab reservations
        conn.execute("DELETE FROM lab_reservations")
        
        # Reset free_sessions_used for all students
        conn.execute("UPDATE students SET free_sessions_used = 0")
        
        # Log the action
        conn.execute("""
            INSERT INTO admin_logs (admin_id, action, timestamp)
            VALUES (?, 'Reset all student data (sessions, points, and reservations)', datetime('now'))
        """, (session['user'],))
        
        conn.commit()
        conn.close()
        
        flash("All student data (sessions, behavior points, and reservations) has been reset.", "success")
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
    
    return redirect(url_for('admin_system_management'))

@app.route('/admin/system')
def admin_system_management():
    """Admin system management page."""
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    return render_template('admin_system_management.html')

@app.route('/admin/backup_database')
def admin_backup_database():
    """Download a backup of the database."""
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    try:
        # Create a copy of the database file
        import os
        import shutil
        from datetime import datetime
        import io
        
        # Get the database file path
        db_path = 'users.db'
        
        # Create a backup file name with timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        backup_filename = f"backup_{timestamp}.db"
        
        # Read the database file into memory
        with open(db_path, 'rb') as f:
            db_data = f.read()
        
        # Create a memory file-like object
        memory_file = io.BytesIO(db_data)
        memory_file.seek(0)
        
        # Log the action
        conn = get_db_connection()
        conn.execute("""
            INSERT INTO admin_logs (admin_id, action, timestamp)
            VALUES (?, 'Database backup created', datetime('now'))
        """, (session['user'],))
        conn.commit()
        conn.close()
        
        # Send the file as download attachment
        return send_file(
            memory_file,
            mimetype='application/octet-stream',
            as_attachment=True,
            download_name=backup_filename
        )
    except Exception as e:
        flash(f"Error creating database backup: {str(e)}", "danger")
        return redirect(url_for('admin_system_management'))

@app.route('/admin/restore_database', methods=['POST'])
def admin_restore_database():
    """Restore the database from a backup file."""
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    if 'database_file' not in request.files:
        flash("No file selected for upload.", "warning")
        return redirect(url_for('admin_system_management'))
    
    file = request.files['database_file']
    
    if file.filename == '':
        flash("No file selected for upload.", "warning")
        return redirect(url_for('admin_system_management'))
    
    try:
        import os
        import shutil
        from datetime import datetime
        
        # Get the database file path
        db_path = 'users.db'
        
        # Create a backup of the current database before replacing it
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = f"auto_backup_before_restore_{timestamp}.db"
        shutil.copy2(db_path, backup_path)
        
        # Save the uploaded file as the new database
        file.save(db_path)
        
        # No need to log in the restored database as it might not have the same schema
        # but we can try
        try:
            conn = get_db_connection()
            conn.execute("""
                INSERT INTO admin_logs (admin_id, action, timestamp)
                VALUES (?, 'Database restored from backup', datetime('now'))
            """, (session['user'],))
            conn.commit()
            conn.close()
        except:
            # If this fails, it's not critical - the restore already happened
            pass
        
        flash("Database successfully restored from backup file.", "success")
    except Exception as e:
        flash(f"Error restoring database: {str(e)}", "danger")
    
    return redirect(url_for('admin_system_management'))

@app.route('/api/search_student', methods=['POST'])
def search_student():
    """API endpoint to search for a student by ID or name."""
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Access denied. Admin privileges required.'}), 403
    
    search_query = request.json.get('query', '').strip()
    
    if not search_query:
        return jsonify({'success': False, 'message': 'Search query is required'}), 400
    
    try:
        conn = get_db_connection()
        
        # Search by ID or name (case-insensitive)
        students = conn.execute("""
            SELECT * FROM students 
            WHERE idno = ? 
            OR LOWER(firstname || ' ' || lastname) LIKE LOWER(?)
            OR LOWER(lastname || ' ' || firstname) LIKE LOWER(?)
            LIMIT 1
        """, (search_query, f'%{search_query}%', f'%{search_query}%')).fetchall()
        
        if not students:
            return jsonify({'success': False, 'message': 'No student found with the provided ID or name'}), 404
        
        student = dict(students[0])
        
        # Calculate remaining sessions
        if student['course'] in ("BSIT", "BSCS"):
            max_sessions = 30
        else:
            max_sessions = 15
            
        # Count completed lab sessions
        used_sessions = conn.execute("""
            SELECT COUNT(*) as c 
            FROM lab_sessions 
            WHERE student_id = ? AND status = 'Completed'
        """, (student['idno'],)).fetchone()['c']
        
        # Calculate remaining sessions
        remaining_sessions = max_sessions - used_sessions
        if remaining_sessions < 0:
            remaining_sessions = 0
            
        # Calculate behavior points and free sessions
        behavior_points = conn.execute("""
            SELECT COALESCE(SUM(behavior_points), 0) as total
            FROM lab_sessions
            WHERE student_id = ? AND behavior_points > 0
        """, (student['idno'],)).fetchone()['total']
        
        free_sessions = behavior_points // 3
        
        # Add free sessions to remaining sessions
        total_remaining = remaining_sessions + free_sessions
        
        # Add computed fields to student data
        student['remaining_sessions'] = remaining_sessions
        student['behavior_points'] = behavior_points
        student['free_sessions'] = free_sessions
        student['total_remaining'] = total_remaining
        
        conn.close()
        
        return jsonify({
            'success': True,
            'student': student
        })
        
    except sqlite3.Error as e:
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500

@app.route('/admin/reset_student_session', methods=['POST'])
def admin_reset_student_session():
    """Reset sessions for a specific student."""
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Access denied. Admin privileges required.'}), 403
    
    student_id = request.json.get('student_id')
    
    if not student_id:
        return jsonify({'success': False, 'message': 'Student ID is required'}), 400
    
    try:
        conn = get_db_connection()
        
        # Verify the student exists
        student = conn.execute("SELECT * FROM students WHERE idno = ?", (student_id,)).fetchone()
        
        if not student:
            conn.close()
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        
        # Delete all completed lab sessions for this student
        conn.execute("DELETE FROM lab_sessions WHERE student_id = ? AND status = 'Completed'", (student_id,))
        
        # Delete all reservations for this student
        conn.execute("DELETE FROM lab_reservations WHERE student_id = ?", (student_id,))
        
        # Reset free_sessions_used for this student
        conn.execute("UPDATE students SET free_sessions_used = 0 WHERE idno = ?", (student_id,))
        
        # Log the action
        conn.execute("""
            INSERT INTO admin_logs (admin_id, action, timestamp)
            VALUES (?, ?, datetime('now'))
        """, (session['user'], f'Reset sessions and cleared reservations for student {student_id}'))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f"Sessions reset and reservations cleared successfully for student {student_id}"
        })
        
    except sqlite3.Error as e:
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500

@app.route('/admin/lab_resources', methods=['GET', 'POST'])
def admin_lab_resources():
    # Check if user is logged in and is admin
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        
        if request.method == 'POST':
            if 'add_resource' in request.form:
                # Add new resource
                title = request.form.get('title')
                description = request.form.get('description')
                resource_type = request.form.get('resource_type')
                file_path = None
                
                # Handle file upload if provided
                if 'resource_file' in request.files and request.files['resource_file'].filename:
                    file = request.files['resource_file']
                    if file.filename:
                        # Ensure uploads directory exists
                        uploads_dir = os.path.join('static', 'uploads', 'resources')
                        os.makedirs(uploads_dir, exist_ok=True)
                        
                        # Generate unique filename
                        filename = secure_filename(file.filename)
                        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                        unique_filename = f"{timestamp}_{filename}"
                        file_path = os.path.join('uploads', 'resources', unique_filename)
                        
                        # Save the file
                        file.save(os.path.join('static', file_path))
                
                # Insert into database
                conn.execute("""
                    INSERT INTO lab_resources (title, description, file_path, resource_type, posted_by, posted_date, is_active)
                    VALUES (?, ?, ?, ?, ?, datetime('now'), 1)
                """, (title, description, file_path, resource_type, session['user']))
                
                # Log the action
                conn.execute("""
                    INSERT INTO admin_logs (admin_id, action, timestamp)
                    VALUES (?, 'Added new lab resource', datetime('now'))
                """, (session['user'],))
                
                conn.commit()
                flash("Resource added successfully!", "success")
                return redirect(url_for('admin_lab_resources'))
            
            elif 'edit_resource' in request.form:
                # Edit existing resource
                resource_id = request.form.get('resource_id')
                title = request.form.get('title')
                description = request.form.get('description')
                resource_type = request.form.get('resource_type')
                
                # Get existing resource to check if there's a file already
                existing_resource = conn.execute("SELECT file_path FROM lab_resources WHERE resource_id = ?", 
                                                (resource_id,)).fetchone()
                file_path = existing_resource['file_path'] if existing_resource else None
                
                # Handle file upload if provided
                if 'resource_file' in request.files and request.files['resource_file'].filename:
                    file = request.files['resource_file']
                    if file.filename:
                        # Remove old file if exists
                        if file_path and os.path.exists(os.path.join('static', file_path)):
                            os.remove(os.path.join('static', file_path))
                        
                        # Ensure uploads directory exists
                        uploads_dir = os.path.join('static', 'uploads', 'resources')
                        os.makedirs(uploads_dir, exist_ok=True)
                        
                        # Generate unique filename
                        filename = secure_filename(file.filename)
                        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                        unique_filename = f"{timestamp}_{filename}"
                        file_path = os.path.join('uploads', 'resources', unique_filename)
                        
                        # Save the file
                        file.save(os.path.join('static', file_path))
                
                # Update database
                conn.execute("""
                    UPDATE lab_resources 
                    SET title = ?, description = ?, file_path = ?, resource_type = ?
                    WHERE resource_id = ?
                """, (title, description, file_path, resource_type, resource_id))
                
                # Log the action
                conn.execute("""
                    INSERT INTO admin_logs (admin_id, action, timestamp)
                    VALUES (?, 'Updated lab resource', datetime('now'))
                """, (session['user'],))
                
                conn.commit()
                flash("Resource updated successfully!", "success")
                return redirect(url_for('admin_lab_resources'))
        
        # Get all resources
        resources = conn.execute("""
            SELECT * FROM lab_resources
            ORDER BY posted_date DESC
        """).fetchall()
        
        return render_template('admin_lab_resources.html', resources=resources)
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_dashboard'))
    finally:
        if conn:
            conn.close()

@app.route('/admin/lab_resources/toggle/<int:resource_id>', methods=['GET'])
def toggle_resource(resource_id):
    # Check if user is logged in and is admin
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        
        # Get current status
        resource = conn.execute("SELECT is_active FROM lab_resources WHERE resource_id = ?", 
                               (resource_id,)).fetchone()
        
        if not resource:
            flash("Resource not found.", "danger")
            return redirect(url_for('admin_lab_resources'))
        
        # Toggle status
        new_status = 0 if resource['is_active'] else 1
        status_text = "enabled" if new_status else "disabled"
        
        conn.execute("UPDATE lab_resources SET is_active = ? WHERE resource_id = ?", 
                    (new_status, resource_id))
        
        # Log the action
        conn.execute("""
            INSERT INTO admin_logs (admin_id, action, timestamp)
            VALUES (?, ?, datetime('now'))
        """, (session['user'], f"Resource #{resource_id} {status_text}"))
        
        conn.commit()
        flash(f"Resource has been {status_text}.", "success")
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
    finally:
        if conn:
            conn.close()
    
    return redirect(url_for('admin_lab_resources'))

@app.route('/admin/lab_resources/delete/<int:resource_id>', methods=['GET'])
def delete_resource(resource_id):
    # Check if user is logged in and is admin
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        
        # Get resource file path
        resource = conn.execute("SELECT file_path FROM lab_resources WHERE resource_id = ?", 
                               (resource_id,)).fetchone()
        
        if not resource:
            flash("Resource not found.", "danger")
            return redirect(url_for('admin_lab_resources'))
        
        # Delete file if exists
        if resource['file_path'] and os.path.exists(os.path.join('static', resource['file_path'])):
            os.remove(os.path.join('static', resource['file_path']))
        
        # Delete from database
        conn.execute("DELETE FROM lab_resources WHERE resource_id = ?", (resource_id,))
        
        # Log the action
        conn.execute("""
            INSERT INTO admin_logs (admin_id, action, timestamp)
            VALUES (?, ?, datetime('now'))
        """, (session['user'], f"Deleted resource #{resource_id}"))
        
        conn.commit()
        flash("Resource has been deleted.", "success")
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
    finally:
        if conn:
            conn.close()
    
    return redirect(url_for('admin_lab_resources'))

@app.route('/lab_resources')
def lab_resources():
    # Check if user is logged in
    if 'user' not in session:
        flash("Please log in to access lab resources.", "warning")
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        
        # Get only active resources for students
        if session.get('is_admin'):
            # Admins can see all resources
            resources = conn.execute("""
                SELECT * FROM lab_resources
                ORDER BY posted_date DESC
            """).fetchall()
        else:
            # Students can only see active resources
            resources = conn.execute("""
                SELECT * FROM lab_resources
                WHERE is_active = 1
                ORDER BY posted_date DESC
            """).fetchall()
        
        return render_template('lab_resources.html', resources=resources)
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('dashboard'))
    finally:
        if conn:
            conn.close()

@app.route('/lab_resources/view/<int:resource_id>')
def view_resource(resource_id):
    # Check if user is logged in
    if 'user' not in session:
        flash("Please log in to access lab resources.", "warning")
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        
        # Get resource
        resource = conn.execute("""
            SELECT * FROM lab_resources
            WHERE resource_id = ?
        """, (resource_id,)).fetchone()
        
        if not resource:
            flash("Resource not found.", "danger")
            return redirect(url_for('lab_resources'))
        
        # Check if resource is active (students only)
        if not session.get('is_admin') and not resource['is_active']:
            flash("This resource is currently unavailable.", "warning")
            return redirect(url_for('lab_resources'))
        
        # Increment view count
        conn.execute("""
            UPDATE lab_resources
            SET view_count = view_count + 1
            WHERE resource_id = ?
        """, (resource_id,))
        
        conn.commit()
        
        # If the resource has a file path, check the file type
        file_type = None
        if resource['file_path']:
            file_extension = os.path.splitext(resource['file_path'])[1].lower()
            
            # Classify file type
            if file_extension in ['.pdf']:
                file_type = 'pdf'
            elif file_extension in ['.jpg', '.jpeg', '.png', '.gif']:
                file_type = 'image'
            elif file_extension in ['.doc', '.docx', '.ppt', '.pptx', '.xls', '.xlsx']:
                file_type = 'document'
            else:
                file_type = 'other'
        
        return render_template('view_resource.html', resource=resource, file_type=file_type)
        
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('lab_resources'))
    finally:
        if conn:
            conn.close()

@app.route('/lab_resources/download/<int:resource_id>')
def download_resource(resource_id):
    # Check if user is logged in
    if 'user' not in session:
        flash("Please log in to download resources.", "warning")
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        
        # Get resource
        resource = conn.execute("""
            SELECT * FROM lab_resources
            WHERE resource_id = ?
        """, (resource_id,)).fetchone()
        
        if not resource or not resource['file_path']:
            flash("Resource or file not found.", "danger")
            return redirect(url_for('lab_resources'))
        
        # Check if resource is active (students only)
        if not session.get('is_admin') and not resource['is_active']:
            flash("This resource is currently unavailable.", "warning")
            return redirect(url_for('lab_resources'))
        
        # File path
        file_path = os.path.join('static', resource['file_path'])
        
        if not os.path.exists(file_path):
            flash("File not found on server.", "danger")
            return redirect(url_for('lab_resources'))
        
        # Generate filename for download
        filename = os.path.basename(resource['file_path'])
        if '_' in filename:
            # Remove timestamp prefix from filename
            filename = filename.split('_', 1)[1]
        
        return send_file(file_path, as_attachment=True, download_name=filename)
        
    except Exception as e:
        flash(f"Error downloading file: {str(e)}", "danger")
        return redirect(url_for('lab_resources'))
    finally:
        if conn:
            conn.close()

@app.route('/student/lab-schedule')
def student_lab_schedule():
    # Add session check
    if 'user' not in session:
        flash('Access denied: Student login required', 'danger')
        return redirect(url_for('home'))
    
    # Get all laboratories
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Fetch laboratories - remove software column reference
    cursor.execute('''
        SELECT lab_id, building, room_number, capacity, equipment
        FROM laboratories
        ORDER BY building, room_number
    ''')
    labs = [dict(row) for row in cursor.fetchall()]
    
    # Fetch lab schedules with joined laboratory info - update to match schema
    cursor.execute('''
        SELECT 
            s.schedule_id, 
            l.lab_id,
            l.building,
            l.room_number,
            s.day_of_week as day,
            s.start_time,
            s.end_time,
            s.is_available,
            s.reserved_for
        FROM lab_schedules s
        JOIN laboratories l ON s.lab_id = l.lab_id
        ORDER BY s.day_of_week, s.start_time
    ''')
    schedules = [dict(row) for row in cursor.fetchall()]
    
    # Define time slots with format HH:MM - HH:MM
    time_slots_24h = [
        "07:00 - 08:00", "08:00 - 09:00", "09:00 - 10:00", "10:00 - 11:00",
        "11:00 - 12:00", "12:00 - 13:00", "13:00 - 14:00", "14:00 - 15:00",
        "15:00 - 16:00", "16:00 - 17:00", "17:00 - 18:00", "18:00 - 19:00",
        "19:00 - 20:00", "20:00 - 21:00"
    ]
    
    # Format time slots for display
    time_slots = []
    for slot in time_slots_24h:
        start, end = slot.split(' - ')
        formatted_start = format_time(start)
        formatted_end = format_time(end)
        time_slots.append(f"{formatted_start} - {formatted_end}")
    
    cursor.close()
    conn.close()
    
    return render_template('student_lab_schedule.html', labs=labs, schedules=schedules, time_slots=time_slots, default_lab_id=517)

@app.route('/reserve-schedule', methods=['POST'])
def reserve_schedule():
    # For now, just redirect back to the schedule page with a message
    flash('Reservation functionality will be implemented soon.', 'info')
    return redirect(url_for('student_lab_schedule'))

@app.route('/admin/lab-schedule')
def admin_lab_schedule():
    # Check if user is logged in and is admin
    if 'user' not in session or session.get('is_admin') != True:
        flash('Access denied: Admin login required', 'danger')
        return redirect(url_for('home'))
    
    # Get all laboratories
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Create the tables if they don't exist
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS laboratories (
            lab_id INTEGER PRIMARY KEY AUTOINCREMENT,
            building TEXT NOT NULL,
            room_number TEXT NOT NULL,
            capacity INTEGER NOT NULL,
            equipment TEXT,
            UNIQUE(building, room_number)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS lab_schedules (
            schedule_id INTEGER PRIMARY KEY AUTOINCREMENT,
            lab_id INTEGER NOT NULL,
            day_of_week TEXT NOT NULL CHECK (day_of_week IN ('Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday')),
            start_time TIME NOT NULL,
            end_time TIME NOT NULL,
            is_available BOOLEAN DEFAULT 1,
            reserved_for TEXT,
            FOREIGN KEY (lab_id) REFERENCES laboratories(lab_id),
            UNIQUE(lab_id, day_of_week, start_time)
        )
    ''')
    
    conn.commit()
    
    # Fetch laboratories - remove software column
    cursor.execute('''
        SELECT lab_id, building, room_number, capacity, equipment
        FROM laboratories
        ORDER BY building, room_number
    ''')
    labs = [dict(row) for row in cursor.fetchall()]
    
    # Fetch lab schedules with joined laboratory info - update to match schema
    cursor.execute('''
        SELECT 
            s.schedule_id, 
            l.lab_id,
            l.building,
            l.room_number,
            s.day_of_week as day,
            s.start_time,
            s.end_time,
            s.is_available,
            s.reserved_for
        FROM lab_schedules s
        JOIN laboratories l ON s.lab_id = l.lab_id
        ORDER BY s.day_of_week, s.start_time
    ''')
    schedules = [dict(row) for row in cursor.fetchall()]
    
    # Define time slots with format HH:MM - HH:MM
    time_slots_24h = [
        "07:00 - 08:00", "08:00 - 09:00", "09:00 - 10:00", "10:00 - 11:00",
        "11:00 - 12:00", "12:00 - 13:00", "13:00 - 14:00", "14:00 - 15:00",
        "15:00 - 16:00", "16:00 - 17:00", "17:00 - 18:00", "18:00 - 19:00",
        "19:00 - 20:00", "20:00 - 21:00"
    ]
    
    # Format time slots for display
    time_slots = []
    for slot in time_slots_24h:
        start, end = slot.split(' - ')
        formatted_start = format_time(start)
        formatted_end = format_time(end)
        time_slots.append(f"{formatted_start} - {formatted_end}")
    
    cursor.close()
    conn.close()
    
    return render_template('admin_lab_schedule.html', labs=labs, schedules=schedules, time_slots=time_slots, default_lab_id=517)

@app.route('/add_schedule', methods=['POST'])
def add_schedule():
    # Check if user is logged in and is admin
    if 'user' not in session or session.get('is_admin') != True:
        flash('Access denied: Admin login required', 'danger')
        return redirect(url_for('home'))
    
    if request.method == 'POST':
        lab_id = request.form.get('lab_id')
        day = request.form.get('day')
        start_time = request.form.get('start_time')
        end_time = request.form.get('end_time')
        status = request.form.get('status')
        reserved_for = request.form.get('reserved_for', '')
        
        # Validate inputs
        if not all([lab_id, day, start_time, end_time, status]):
            flash('All fields are required', 'danger')
            return redirect(url_for('admin_lab_schedule'))
        
        # Check if end time is after start time
        if start_time >= end_time:
            flash('End time must be after start time', 'danger')
            return redirect(url_for('admin_lab_schedule'))
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Check for scheduling conflicts
            cursor.execute('''
                SELECT * FROM lab_schedules 
                WHERE lab_id = ? AND day_of_week = ? AND 
                ((start_time <= ? AND end_time > ?) OR
                 (start_time < ? AND end_time >= ?) OR
                 (start_time >= ? AND end_time <= ?))
            ''', (lab_id, day, start_time, start_time, end_time, end_time, start_time, end_time))
            
            if cursor.fetchone():
                flash('Schedule conflict detected. Another schedule exists for this time slot.', 'danger')
                return redirect(url_for('admin_lab_schedule'))
            
            # Insert new schedule
            cursor.execute('''
                INSERT INTO lab_schedules (lab_id, day_of_week, start_time, end_time, is_available, reserved_for)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (lab_id, day, start_time, end_time, 1 if status == 'Available' else 0, reserved_for))
            
            conn.commit()
            flash('Schedule added successfully', 'success')
            
        except Exception as e:
            flash(f'Error adding schedule: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
            
    return redirect(url_for('admin_lab_schedule'))

@app.route('/edit_schedule/<int:schedule_id>', methods=['GET', 'POST'])
def edit_schedule(schedule_id):
    # Check if user is logged in and is admin
    if 'user' not in session or session.get('is_admin') != True:
        flash('Access denied: Admin login required', 'danger')
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        lab_id = request.form.get('lab_id')
        day_of_week = request.form.get('day')
        start_time = request.form.get('start_time')
        end_time = request.form.get('end_time')
        status = request.form.get('status')
        reserved_for = request.form.get('reserved_for', '')
        
        # Validate inputs
        if not all([lab_id, day_of_week, start_time, end_time, status]):
            flash('All fields are required', 'danger')
            return redirect(url_for('edit_schedule', schedule_id=schedule_id))
        
        # Check if end time is after start time
        if start_time >= end_time:
            flash('End time must be after start time', 'danger')
            return redirect(url_for('edit_schedule', schedule_id=schedule_id))
        
        try:
            # Check for scheduling conflicts, excluding current schedule
            cursor.execute('''
                SELECT * FROM lab_schedules 
                WHERE lab_id = ? AND day_of_week = ? AND schedule_id != ? AND
                ((start_time <= ? AND end_time > ?) OR
                 (start_time < ? AND end_time >= ?) OR
                 (start_time >= ? AND end_time <= ?))
            ''', (lab_id, day_of_week, schedule_id, start_time, start_time, end_time, end_time, start_time, end_time))
            
            if cursor.fetchone():
                flash('Schedule conflict detected. Another schedule exists for this time slot.', 'danger')
                return redirect(url_for('edit_schedule', schedule_id=schedule_id))
            
            # Update schedule
            cursor.execute('''
                UPDATE lab_schedules 
                SET lab_id = ?, day_of_week = ?, start_time = ?, end_time = ?, is_available = ?, reserved_for = ?
                WHERE schedule_id = ?
            ''', (lab_id, day_of_week, start_time, end_time, 1 if status == 'Available' else 0, reserved_for, schedule_id))
            
            conn.commit()
            flash('Schedule updated successfully', 'success')
            return redirect(url_for('admin_lab_schedule'))
            
        except Exception as e:
            flash(f'Error updating schedule: {str(e)}', 'danger')
            return redirect(url_for('edit_schedule', schedule_id=schedule_id))
    
    # GET request - Display edit form
    try:
        # Get schedule details
        cursor.execute('''
            SELECT 
                schedule_id, 
                lab_id,
                day_of_week,
                start_time,
                end_time,
                is_available,
                reserved_for
            FROM lab_schedules 
            WHERE schedule_id = ?
        ''', (schedule_id,))
        schedule = cursor.fetchone()
        
        if not schedule:
            flash('Schedule not found', 'danger')
            return redirect(url_for('admin_lab_schedule'))
        
        # Get all laboratories for the select dropdown
        cursor.execute('''
            SELECT lab_id, building, room_number
            FROM laboratories
            ORDER BY building, room_number
        ''')
        labs = [dict(row) for row in cursor.fetchall()]
        
        schedule_dict = dict(schedule)
        
        # Add a 'day' key for compatibility with the template
        schedule_dict['day'] = schedule_dict['day_of_week']
        
        # Add a 'status' key for compatibility with the template
        schedule_dict['status'] = 'Available' if schedule_dict['is_available'] else 'Unavailable'
        
        return render_template('edit_schedule.html', schedule=schedule_dict, labs=labs)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('admin_lab_schedule'))
    finally:
        cursor.close()
        conn.close()

@app.route('/delete_schedule/<int:schedule_id>', methods=['POST'])
def delete_schedule(schedule_id):
    # Check if user is logged in and is admin
    if 'user' not in session or session.get('is_admin') != True:
        flash('Access denied: Admin login required', 'danger')
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if schedule exists
        cursor.execute('SELECT schedule_id FROM lab_schedules WHERE schedule_id = ?', (schedule_id,))
        if not cursor.fetchone():
            flash('Schedule not found', 'danger')
            return redirect(url_for('admin_lab_schedule'))
        
        # Delete schedule
        cursor.execute('DELETE FROM lab_schedules WHERE schedule_id = ?', (schedule_id,))
        conn.commit()
        
        flash('Schedule deleted successfully', 'success')
        return redirect(url_for('admin_lab_schedule'))
    except Exception as e:
        flash(f'Error deleting schedule: {str(e)}', 'danger')
        return redirect(url_for('admin_lab_schedule'))
    finally:
        cursor.close()
        conn.close()

@app.route('/add_laboratory', methods=['POST'])
def add_laboratory():
    # Check if user is logged in and is admin
    if 'user' not in session or session.get('is_admin') != True:
        flash('Access denied: Admin login required', 'danger')
        return redirect(url_for('home'))
    
    if request.method == 'POST':
        building = request.form.get('building')
        room_number = request.form.get('room_number')
        capacity = request.form.get('capacity')
        equipment = request.form.get('equipment', '')
        
        # Validate inputs
        if not all([building, room_number, capacity]):
            flash('Building, room number, and capacity are required', 'danger')
            return redirect(url_for('admin_lab_schedule'))
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Check if laboratory already exists
            cursor.execute('''
                SELECT * FROM laboratories 
                WHERE building = ? AND room_number = ?
            ''', (building, room_number))
            
            if cursor.fetchone():
                flash('Laboratory already exists', 'danger')
                return redirect(url_for('admin_lab_schedule'))
            
            # Insert new laboratory - remove software
            cursor.execute('''
                INSERT INTO laboratories (building, room_number, capacity, equipment)
                VALUES (?, ?, ?, ?)
            ''', (building, room_number, capacity, equipment))
            
            conn.commit()
            flash('Laboratory added successfully', 'success')
            
        except Exception as e:
            flash(f'Error adding laboratory: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
            
    return redirect(url_for('admin_lab_schedule'))

@app.route('/edit_laboratory/<int:lab_id>', methods=['GET', 'POST'])
def edit_laboratory(lab_id):
    # Check if user is logged in and is admin
    if 'user' not in session or session.get('is_admin') != True:
        flash('Access denied: Admin login required', 'danger')
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        building = request.form.get('building')
        room_number = request.form.get('room_number')
        capacity = request.form.get('capacity')
        equipment = request.form.get('equipment', '')
        
        # Validate inputs
        if not all([building, room_number, capacity]):
            flash('Building, room number, and capacity are required', 'danger')
            return redirect(url_for('edit_laboratory', lab_id=lab_id))
        
        try:
            # Check if another laboratory with same building/room exists
            cursor.execute('''
                SELECT * FROM laboratories 
                WHERE building = ? AND room_number = ? AND lab_id != ?
            ''', (building, room_number, lab_id))
            
            if cursor.fetchone():
                flash('Another laboratory with this building and room number already exists', 'danger')
                return redirect(url_for('edit_laboratory', lab_id=lab_id))
            
            # Update laboratory - remove software
            cursor.execute('''
                UPDATE laboratories 
                SET building = ?, room_number = ?, capacity = ?, equipment = ?
                WHERE lab_id = ?
            ''', (building, room_number, capacity, equipment, lab_id))
            
            conn.commit()
            flash('Laboratory updated successfully', 'success')
            return redirect(url_for('admin_lab_schedule'))
            
        except Exception as e:
            flash(f'Error updating laboratory: {str(e)}', 'danger')
            return redirect(url_for('edit_laboratory', lab_id=lab_id))
    
    # GET request - Display edit form
    try:
        cursor.execute('SELECT * FROM laboratories WHERE lab_id = ?', (lab_id,))
        lab = cursor.fetchone()
        
        if not lab:
            flash('Laboratory not found', 'danger')
            return redirect(url_for('admin_lab_schedule'))
        
        return render_template('edit_laboratory.html', lab=dict(lab))
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('admin_lab_schedule'))
    finally:
        cursor.close()
        conn.close()

@app.route('/delete_laboratory/<int:lab_id>', methods=['POST'])
def delete_laboratory(lab_id):
    # Check if user is logged in and is admin
    if 'user' not in session or session.get('is_admin') != True:
        flash('Access denied: Admin login required', 'danger')
        return redirect(url_for('home'))
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if laboratory exists
        cursor.execute('SELECT lab_id FROM laboratories WHERE lab_id = ?', (lab_id,))
        if not cursor.fetchone():
            flash('Laboratory not found', 'danger')
            return redirect(url_for('admin_lab_schedule'))
        
        # Check if there are schedules using this laboratory
        cursor.execute('SELECT schedule_id FROM lab_schedules WHERE lab_id = ?', (lab_id,))
        schedules = cursor.fetchall()
        
        # Begin transaction
        conn.execute('BEGIN TRANSACTION')
        
        # Delete all related schedules
        if schedules:
            schedule_ids = [s[0] for s in schedules]
            
            for schedule_id in schedule_ids:
                # Delete the schedule
                cursor.execute('DELETE FROM lab_schedules WHERE schedule_id = ?', (schedule_id,))
        
        # Delete the laboratory
        cursor.execute('DELETE FROM laboratories WHERE lab_id = ?', (lab_id,))
        
        # Commit transaction
        conn.commit()
        
        flash('Laboratory and associated schedules deleted successfully', 'success')
        return redirect(url_for('admin_lab_schedule'))
    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error deleting laboratory: {str(e)}', 'danger')
        return redirect(url_for('admin_lab_schedule'))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/upload_schedule', methods=['POST'])
def upload_schedule():
    # Check if user is logged in and is admin
    if 'user' not in session or session.get('is_admin') != True:
        flash('Access denied: Admin login required', 'danger')
        return redirect(url_for('home'))
    
    if 'schedule_file' not in request.files:
        flash('No file part', 'danger')
        return redirect(url_for('admin_lab_schedule'))
    
    file = request.files['schedule_file']
    if file.filename == '':
        flash('No selected file', 'danger')
        return redirect(url_for('admin_lab_schedule'))
    
    if file and file.filename.endswith('.csv'):
        try:
            # Read the CSV file
            csv_data = file.read().decode('utf-8')
            csv_reader = csv.reader(csv_data.splitlines())
            
            # Skip header row
            next(csv_reader)
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            success_count = 0
            error_count = 0
            
            for row in csv_reader:
                if len(row) < 5:  # lab_id, day, start_time, end_time, status
                    error_count += 1
                    continue
                
                lab_id, day, start_time, end_time, status = row
                
                # Validate day
                valid_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
                if day not in valid_days:
                    error_count += 1
                    continue
                
                # Validate times
                try:
                    # Ensure valid time format
                    datetime.strptime(start_time, '%H:%M')
                    datetime.strptime(end_time, '%H:%M')
                    
                    # Check if end time is after start time
                    if start_time >= end_time:
                        error_count += 1
                        continue
                except ValueError:
                    error_count += 1
                    continue
                
                # Validate status
                valid_statuses = ['Available', 'Unavailable']
                if status not in valid_statuses:
                    error_count += 1
                    continue
                
                # Check if laboratory exists
                cursor.execute('SELECT id FROM laboratories WHERE id = ?', (lab_id,))
                if not cursor.fetchone():
                    error_count += 1
                    continue
                
                # Check for scheduling conflicts
                cursor.execute('''
                    SELECT * FROM lab_schedules 
                    WHERE lab_id = ? AND day = ? AND 
                    ((start_time <= ? AND end_time > ?) OR
                     (start_time < ? AND end_time >= ?) OR
                     (start_time >= ? AND end_time <= ?))
                ''', (lab_id, day, start_time, start_time, end_time, end_time, start_time, end_time))
                
                if cursor.fetchone():
                    error_count += 1
                    continue
                
                # Insert new schedule
                cursor.execute('''
                    INSERT INTO lab_schedules (lab_id, day, start_time, end_time, status)
                    VALUES (?, ?, ?, ?, ?)
                ''', (lab_id, day, start_time, end_time, status))
                
                success_count += 1
            
            conn.commit()
            
            if success_count > 0:
                flash(f'Successfully imported {success_count} schedules', 'success')
            if error_count > 0:
                flash(f'Failed to import {error_count} schedules due to validation errors', 'warning')
            
        except Exception as e:
            flash(f'Error importing schedules: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
    else:
        flash('File must be a CSV', 'danger')
    
    return redirect(url_for('admin_lab_schedule'))

@app.route('/download_schedule_template')
def download_schedule_template():
    # Check if user is logged in and is admin
    if 'user' not in session or session.get('is_admin') != True:
        flash('Access denied: Admin login required', 'danger')
        return redirect(url_for('home'))
    
    # Create a CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['lab_id', 'day_of_week', 'start_time', 'end_time', 'is_available', 'reserved_for'])
    
    # Write example rows
    writer.writerow(['1', 'Monday', '08:00', '10:00', '1', ''])
    writer.writerow(['1', 'Monday', '10:00', '12:00', '1', 'BSIT 3'])
    writer.writerow(['2', 'Tuesday', '13:00', '15:00', '0', 'Faculty'])
    
    # Prepare response
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=schedule_template.csv"
    response.headers["content-type"] = "text/csv"
    
    return response

@app.route('/admin/reservations')
def admin_reservations():
    if 'user' not in session or not session.get('is_admin'):
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('home'))
    try:
        conn = get_db_connection()
        # Fetch only pending reservations with student and lab details
        reservations = conn.execute('''
            SELECT lr.reservation_id, lr.student_id, s.lastname, s.firstname, s.midname, s.course, s.year_level,
                   lr.lab_id, l.building, l.room_number, lr.reservation_date, lr.start_time, lr.end_time,
                   lr.purpose, lr.status, lr.created_at, lr.approved_by, lr.approval_date, lr.rejection_reason,
                   lr.computer_number
            FROM lab_reservations lr
            JOIN students s ON lr.student_id = s.idno
            JOIN laboratories l ON lr.lab_id = l.lab_id
            WHERE lr.status = 'Pending'
            ORDER BY lr.reservation_date DESC, lr.start_time DESC
        ''').fetchall()
        
        # Fetch processed reservations (approved/rejected) for logs
        processed_reservations = conn.execute('''
            SELECT lr.reservation_id, lr.student_id, s.lastname, s.firstname, s.midname,
                   lr.lab_id, l.building, l.room_number, lr.reservation_date, lr.start_time, lr.end_time,
                   lr.purpose, lr.status, lr.created_at, lr.approved_by, lr.approval_date, lr.rejection_reason,
                   lr.computer_number
            FROM lab_reservations lr
            JOIN students s ON lr.student_id = s.idno
            JOIN laboratories l ON lr.lab_id = l.lab_id
            WHERE lr.status IN ('Approved', 'Rejected')
            ORDER BY lr.approval_date DESC
            LIMIT 20
        ''').fetchall()
        
        # Fetch all labs with capacity
        labs = conn.execute('''
            SELECT lab_id, building, room_number, capacity
            FROM laboratories
            ORDER BY building, room_number
        ''').fetchall()
        conn.close()
        return render_template('admin_reservations.html', reservations=reservations, processed_reservations=processed_reservations, labs=labs)
    except sqlite3.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for('admin_dashboard'))

@app.route('/api/get_lab_pc_status/<int:lab_id>')
def get_lab_pc_status(lab_id):
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({"error": "Access denied"}), 403
    
    try:
        conn = get_db_connection()
        
        # Check if lab exists
        lab = conn.execute("SELECT * FROM laboratories WHERE lab_id = ?", (lab_id,)).fetchone()
        if not lab:
            conn.close()
            return jsonify({"error": "Laboratory not found"}), 404
        
        # Get PCs for this lab
        pc_status = []
        
        # First, check if we have PCs stored in the database
        lab_pcs = conn.execute("""
            SELECT id, pc_number, status 
            FROM lab_computers 
            WHERE lab_id = ? 
            ORDER BY pc_number
        """, (lab_id,)).fetchall()
        
        # If no PCs found, create default status based on lab capacity
        if not lab_pcs:
            for i in range(1, lab['capacity'] + 1):
                pc_status.append({
                    "pc_number": i,
                    "status": "Available"
                })
        else:
            for pc in lab_pcs:
                pc_status.append({
                    "id": pc['id'],
                    "pc_number": pc['pc_number'],
                    "status": pc['status']
                })
        
        conn.close()
        return jsonify({"lab_id": lab_id, "capacity": lab['capacity'], "pc_status": pc_status})
    
    except sqlite3.Error as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500

@app.route('/api/update_lab_pc_status', methods=['POST'])
def update_lab_pc_status():
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({"error": "Access denied"}), 403
    
    try:
        data = request.json
        lab_id = data.get('lab_id')
        pc_status = data.get('pc_status')
        
        if not lab_id or not pc_status:
            return jsonify({"error": "Missing required data"}), 400
        
        conn = get_db_connection()
        
        # First, check if the lab_computers table exists
        table_exists = conn.execute("""
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name='lab_computers'
        """).fetchone()
        
        # If table doesn't exist, create it
        if not table_exists:
            conn.execute("""
                CREATE TABLE lab_computers (
                    id INTEGER PRIMARY KEY,
                    lab_id INTEGER NOT NULL,
                    pc_number INTEGER NOT NULL,
                    status TEXT DEFAULT 'Available',
                    last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (lab_id) REFERENCES laboratories(lab_id),
                    UNIQUE(lab_id, pc_number)
                )
            """)
        
        # Clear existing PC status for this lab and insert new ones
        conn.execute("DELETE FROM lab_computers WHERE lab_id = ?", (lab_id,))
        
        for pc in pc_status:
            conn.execute("""
                INSERT INTO lab_computers (lab_id, pc_number, status, last_updated)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            """, (lab_id, pc["pc"], pc["status"]))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "PC status updated successfully"})
    
    except sqlite3.Error as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": f"Server error: {str(e)}"}), 500

@app.route('/api/approve_reservation', methods=['POST'])
def approve_reservation():
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({"error": "Access denied"}), 403
    
    data = request.json
    reservation_id = data.get('reservation_id')
    manual_pc_number = data.get('pc_number')  # Optional manual PC assignment

    if not reservation_id:
        return jsonify({"error": "Missing reservation ID"}), 400
    
    try:
        conn = get_db_connection()
        
        # Get reservation details
        reservation = conn.execute('''
            SELECT r.*, l.capacity
            FROM lab_reservations r
            JOIN laboratories l ON r.lab_id = l.lab_id
            WHERE r.reservation_id = ?
        ''', (reservation_id,)).fetchone()
        
        if not reservation:
            conn.close()
            return jsonify({"error": "Reservation not found"}), 404
        
        if reservation['status'] != 'Pending':
            conn.close()
            return jsonify({"error": "Only pending reservations can be approved"}), 400
        
        # Auto-assign a computer if not manually specified
        pc_number = manual_pc_number
        if not pc_number:
            # Check available computers for this lab on reservation date and time
            available_pcs = conn.execute('''
                SELECT pc_number, status
                FROM lab_computers
                WHERE lab_id = ? AND status = 'Available'
                ORDER BY pc_number
            ''', (reservation['lab_id'],)).fetchall()
            
            # If there are no computers tracked yet, initialize them
            if not available_pcs and reservation['capacity'] > 0:
                # Create available computers for this lab if none exist
                for i in range(1, reservation['capacity'] + 1):
                    conn.execute('''
                        INSERT INTO lab_computers (lab_id, pc_number, status)
                        VALUES (?, ?, 'Available')
                    ''', (reservation['lab_id'], i))
                conn.commit()
                
                # Fetch the newly created computers
                available_pcs = conn.execute('''
                    SELECT pc_number, status
                    FROM lab_computers
                    WHERE lab_id = ? AND status = 'Available'
                    ORDER BY pc_number
                ''', (reservation['lab_id'],)).fetchall()
            
            # Check for existing reservations at the same time to avoid double-booking
            existing_reservations = conn.execute('''
                SELECT computer_number
                FROM lab_reservations
                WHERE lab_id = ? 
                AND reservation_date = ? 
                AND status = 'Approved'
                AND computer_number IS NOT NULL
                AND (
                    (start_time <= ? AND end_time > ?) OR
                    (start_time < ? AND end_time >= ?) OR
                    (start_time >= ? AND end_time <= ?)
                )
            ''', (
                reservation['lab_id'],
                reservation['reservation_date'],
                reservation['start_time'], reservation['start_time'],
                reservation['end_time'], reservation['end_time'],
                reservation['start_time'], reservation['end_time']
            )).fetchall()
            
            # Get list of already booked computers
            booked_computers = [r['computer_number'] for r in existing_reservations]
            
            # Find an available computer not already booked
            assigned_pc = None
            for pc in available_pcs:
                if pc['pc_number'] not in booked_computers:
                    assigned_pc = pc['pc_number']
                    break
            
            if not assigned_pc and available_pcs:
                # If all PCs are booked at this timeslot, just assign the first available
                assigned_pc = available_pcs[0]['pc_number']
            
            pc_number = assigned_pc
        
        if not pc_number:
            conn.close()
            return jsonify({
                "error": "No computers available for assignment. Please manually assign a computer."
            }), 400
        
        # Update reservation status and assign PC
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute('''
            UPDATE lab_reservations
            SET status = 'Approved',
                computer_number = ?,
                approved_by = ?,
                approval_date = ?
            WHERE reservation_id = ?
        ''', (pc_number, session['user'], current_time, reservation_id))
        
        # Update the PC status to "Used" in the lab_computers table
        # This ensures the PC shows as red in the Computer Control panel
        pc_exists = conn.execute('''
            SELECT COUNT(*) as count FROM lab_computers 
            WHERE lab_id = ? AND pc_number = ?
        ''', (reservation['lab_id'], pc_number)).fetchone()['count']
        
        if pc_exists > 0:
            # Update existing PC record
            conn.execute('''
                UPDATE lab_computers
                SET status = 'Used',
                    last_updated = CURRENT_TIMESTAMP
                WHERE lab_id = ? AND pc_number = ?
            ''', (reservation['lab_id'], pc_number))
        else:
            # Create a new PC record if it doesn't exist
            conn.execute('''
                INSERT INTO lab_computers (lab_id, pc_number, status, last_updated)
                VALUES (?, ?, 'Used', CURRENT_TIMESTAMP)
            ''', (reservation['lab_id'], pc_number))
        
        # Create an active lab session for the student
        # This ensures the reservation appears in the Current Sit-In module
        
        # Use current time for check-in instead of reservation time
        current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Create the lab session
        conn.execute('''
            INSERT INTO lab_sessions (
                student_id, lab_id, reservation_id, check_in_time, 
                status, notes, created_by
            ) VALUES (?, ?, ?, ?, 'Active', ?, ?)
        ''', (
            reservation['student_id'], 
            reservation['lab_id'], 
            reservation_id, 
            current_datetime,
            reservation['purpose'],  # Just use the purpose directly without the "Reservation #" prefix
            session['user']
        ))
        
        conn.commit()
        
        # Get updated reservation for response
        updated_reservation = conn.execute('''
            SELECT r.*, s.firstname, s.lastname, l.building, l.room_number
            FROM lab_reservations r
            JOIN students s ON r.student_id = s.idno
            JOIN laboratories l ON r.lab_id = l.lab_id
            WHERE r.reservation_id = ?
        ''', (reservation_id,)).fetchone()
        
        if updated_reservation:
            reservation_data = dict(updated_reservation)
            
            # Add a log entry for the approval
            log_message = f"Reservation #{reservation_id} for {reservation_data['firstname']} {reservation_data['lastname']} approved, PC #{pc_number} assigned"
            conn.execute('''
                INSERT INTO admin_logs (admin_id, action, timestamp)
                VALUES (?, ?, datetime('now'))
            ''', (session['user'], log_message))
            conn.commit()
        
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "Reservation approved successfully",
            "pc_number": pc_number,
            "reservation": dict(updated_reservation) if updated_reservation else None
        })
        
    except sqlite3.Error as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": f"Server error: {str(e)}"}), 500

@app.route('/api/reject_reservation', methods=['POST'])
def reject_reservation():
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({"error": "Access denied"}), 403
    
    data = request.json
    reservation_id = data.get('reservation_id')
    rejection_reason = data.get('rejection_reason', 'No reason provided')

    if not reservation_id:
        return jsonify({"error": "Missing reservation ID"}), 400
    
    try:
        conn = get_db_connection()
        
        # Get reservation details
        reservation = conn.execute('''
            SELECT r.*, s.firstname, s.lastname
            FROM lab_reservations r
            JOIN students s ON r.student_id = s.idno
            WHERE r.reservation_id = ?
        ''', (reservation_id,)).fetchone()
        
        if not reservation:
            conn.close()
            return jsonify({"error": "Reservation not found"}), 404
        
        if reservation['status'] != 'Pending':
            conn.close()
            return jsonify({"error": "Only pending reservations can be rejected"}), 400
        
        # Update reservation status
        conn.execute('''
            UPDATE lab_reservations
            SET status = 'Rejected',
                rejection_reason = ?
            WHERE reservation_id = ?
        ''', (rejection_reason, reservation_id))
        
        # Add a log entry
        log_message = f"Reservation #{reservation_id} for {reservation['firstname']} {reservation['lastname']} rejected: {rejection_reason}"
        conn.execute('''
            INSERT INTO admin_logs (admin_id, action, timestamp)
            VALUES (?, ?, datetime('now'))
        ''', (session['user'], log_message))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "Reservation rejected successfully"
        })
        
    except sqlite3.Error as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": f"Server error: {str(e)}"}), 500

@app.route('/api/get_booked_pcs')
def get_booked_pcs():
    """API endpoint to get PCs that are already booked for a specific time slot"""
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({"error": "Access denied"}), 403
    
    lab_id = request.args.get('lab_id')
    date = request.args.get('date')
    start_time = request.args.get('start_time')
    end_time = request.args.get('end_time')
    
    if not all([lab_id, date, start_time, end_time]):
        return jsonify({"error": "Missing required parameters"}), 400
    
    try:
        conn = get_db_connection()
        
        # Get booked PCs for this time slot
        booked_pcs = conn.execute("""
            SELECT computer_number 
            FROM lab_reservations
            WHERE lab_id = ? 
            AND reservation_date = ? 
            AND status = 'Approved'
            AND computer_number IS NOT NULL
            AND (
                (start_time <= ? AND end_time > ?) OR
                (start_time < ? AND end_time >= ?) OR
                (start_time >= ? AND end_time <= ?)
            )
        """, (lab_id, date, start_time, start_time, end_time, end_time, start_time, end_time)).fetchall()
        
        # Extract PC numbers
        booked_pc_numbers = [pc['computer_number'] for pc in booked_pcs]
        
        # Get computers marked as "Used" in the lab_computers table
        used_pcs = conn.execute("""
            SELECT pc_number 
            FROM lab_computers 
            WHERE lab_id = ? AND status = 'Used'
        """, (lab_id,)).fetchall()
        
        # Combine both lists
        used_pc_numbers = [pc['pc_number'] for pc in used_pcs]
        all_unavailable = list(set(booked_pc_numbers + used_pc_numbers))
        
        conn.close()
        
        return jsonify({
            "lab_id": lab_id,
            "date": date,
            "start_time": start_time,
            "end_time": end_time,
            "booked_pcs": all_unavailable
        })
        
    except sqlite3.Error as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500

@app.route('/api/get_available_pcs')
def get_available_pcs():
    """API endpoint to get PCs that are available for a specific lab and date"""
    lab_id = request.args.get('lab_id')
    reservation_date = request.args.get('date')
    start_time = request.args.get('start_time')
    end_time = request.args.get('end_time')
    
    if not all([lab_id, reservation_date]):
        return jsonify({"error": "Missing required parameters"}), 400
    
    try:
        conn = get_db_connection()
        
        # Check if lab exists
        lab = conn.execute("SELECT * FROM laboratories WHERE lab_id = ?", (lab_id,)).fetchone()
        if not lab:
            conn.close()
            return jsonify({"error": "Laboratory not found"}), 404
        
        # Check if reservation date is today
        is_today = reservation_date == datetime.now().strftime('%Y-%m-%d')
        
        # Get all PCs for this lab
        pc_status = []
        
        # Get PC status for today from lab_computers table if the date is today
        if is_today:
            lab_pcs = conn.execute("""
                SELECT pc_number, status 
                FROM lab_computers 
                WHERE lab_id = ? 
                ORDER BY pc_number
            """, (lab_id,)).fetchall()
            
            # If no PCs found in the database, initialize them all as available
            if not lab_pcs and lab['capacity'] > 0:
                for i in range(1, lab['capacity'] + 1):
                    pc_status.append({
                        "pc_number": i,
                        "status": "Available"
                    })
            else:
                for pc in lab_pcs:
                    pc_status.append({
                        "pc_number": pc['pc_number'],
                        "status": pc['status']
                    })
        else:
            # For future dates, initialize all PCs as available
            for i in range(1, lab['capacity'] + 1):
                pc_status.append({
                    "pc_number": i,
                    "status": "Available"
                })
        
        # Get reserved PCs for this timeslot regardless of date
        reserved_pcs = []
        if start_time and end_time:
            booked_pcs = conn.execute("""
                SELECT computer_number 
                FROM lab_reservations
                WHERE lab_id = ? 
                AND reservation_date = ? 
                AND status = 'Approved'
                AND computer_number IS NOT NULL
                AND (
                    (start_time <= ? AND end_time > ?) OR
                    (start_time < ? AND end_time >= ?) OR
                    (start_time >= ? AND end_time <= ?)
                )
            """, (lab_id, reservation_date, start_time, start_time, end_time, end_time, start_time, end_time)).fetchall()
            
            reserved_pcs = [pc['computer_number'] for pc in booked_pcs]
        
        # Mark reserved PCs as unavailable
        for pc in pc_status:
            if pc['pc_number'] in reserved_pcs:
                pc['status'] = 'Used'
        
        conn.close()
        
        # Return all PCs with their availability status
        return jsonify({
            "lab_id": lab_id,
            "capacity": lab['capacity'],
            "pc_status": pc_status,
            "is_today": is_today
        })
        
    except sqlite3.Error as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True)